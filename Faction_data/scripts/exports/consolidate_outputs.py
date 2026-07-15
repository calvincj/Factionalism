"""
Collapses the output sprawl down to exactly 4 files:

  1. Baidu Data - All People.xlsx       (raw-ish: every person's Baidu match info
                                          + all 3557 parsed career episodes)
  2. Wikipedia Data - All People.xlsx   (raw-ish: every person's Wikipedia match
                                          info, ZH + EN, offices, extracts)
  3. Merged Data - All People.xlsx      (one row per person: Baidu + Wikipedia
                                          side by side + a cross-check flag)
  4. Connections Data.xlsx              (the actual factional-overlap analysis:
                                          Summary/By PBSC/Top Connectors/Top Pair
                                          Evidence/Anchors/Manual vs Physical)

Everything else that was in outputs/ (the old MAIN file, the Wikipedia
cross-check file, the network dataset + CSVs, the docx write-up, the old
README) gets moved into outputs/archive/.
"""

from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
FACTION_DATA_DIR = SCRIPT_DIR.parent.parent  # scripts/exports/ -> scripts/ -> Faction_data/
SOURCE_WORKBOOK = FACTION_DATA_DIR / "source" / "Chinese Leadership Database - Baidu enriched working copy.xlsx"
OUTPUTS = FACTION_DATA_DIR / "outputs"
OLD_MAIN = OUTPUTS / "MAIN - PBSC Overlap Analysis.xlsx"
OLD_WIKI = OUTPUTS / "PBSC Wikipedia cross-check.xlsx"


def autosize(ws, max_width=90):
    for column_cells in ws.columns:
        letter = column_cells[0].column_letter
        width = 10
        for cell in column_cells[:200]:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(max_width, len(value) + 2))
        ws.column_dimensions[letter].width = width


def write_sheet(wb, name, headers, rows):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    autosize(ws)
    return ws


def build_baidu_file():
    src = openpyxl.load_workbook(SOURCE_WORKBOOK, read_only=True, data_only=True)

    cc_rows = list(src["20th Central Committee"].iter_rows(values_only=True))
    cc_headers = [str(h or "").strip() for h in cc_rows[0]]
    idx = {h: i for i, h in enumerate(cc_headers) if h}
    people_headers = [
        "Chinese Name", "Pinyin Name", "Title", "Branch of government",
        "Baidu Query Name", "Baidu URL", "Baidu Last Updated", "Baidu Match Status",
        "Baidu Timeline Episode Count", "Baidu Profile Summary (CN)", "Baidu Notes",
    ]
    people_rows = []
    for r in cc_rows[1:]:
        if not r[idx["Chinese Name"]]:
            continue
        people_rows.append([r[idx[h]] if h in idx else "" for h in people_headers])

    ep_rows_raw = list(src["Baidu Career Episodes"].iter_rows(values_only=True))
    ep_headers = list(ep_rows_raw[0])
    ep_rows = ep_rows_raw[1:]

    out = openpyxl.Workbook()
    out.remove(out.active)
    write_sheet(out, "People", people_headers, people_rows)
    write_sheet(out, "Career Episodes", ep_headers, ep_rows)
    out.save(OUTPUTS / "Baidu Data - All People.xlsx")
    src.close()
    print(f"Baidu Data - All People.xlsx: {len(people_rows)} people, {len(ep_rows)} episodes")


def build_wikipedia_file():
    src = openpyxl.load_workbook(OLD_WIKI, data_only=True)
    ws = src["Wikipedia Cross-Check"]
    rows = list(ws.iter_rows(values_only=True))
    all_headers = list(rows[0])

    keep = [
        "Chinese Name", "Pinyin Name", "Title",
        "Wikipedia ZH Status", "Wikipedia ZH Title", "Wikipedia ZH Note",
        "Wikipedia ZH Offices", "Wikipedia ZH Office Count", "Wikipedia ZH Extract",
        "Wikipedia EN Status", "Wikipedia EN Title", "Wikipedia EN Note",
        "Wikipedia EN Offices", "Wikipedia EN Office Count", "Wikipedia EN Extract",
    ]
    col_idx = [all_headers.index(h) for h in keep]
    out_rows = [[r[i] for i in col_idx] for r in rows[1:]]

    out = openpyxl.Workbook()
    out.remove(out.active)
    write_sheet(out, "People", keep, out_rows)
    out.save(OUTPUTS / "Wikipedia Data - All People.xlsx")
    src.close()
    print(f"Wikipedia Data - All People.xlsx: {len(out_rows)} people")


def build_merged_file():
    src = openpyxl.load_workbook(OLD_MAIN, data_only=True)
    ws = src["Wikipedia Coverage (All 219)"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    data_rows = rows[1:]

    out = openpyxl.Workbook()
    out.remove(out.active)
    write_sheet(out, "People", headers, data_rows)
    out.save(OUTPUTS / "Merged Data - All People.xlsx")
    src.close()
    print(f"Merged Data - All People.xlsx: {len(data_rows)} people")


def build_connections_file():
    src = openpyxl.load_workbook(OLD_MAIN, data_only=True)
    keep_sheets = ["Summary", "By PBSC", "Top Connectors", "Top Pair Evidence", "Anchors", "Manual vs Physical"]

    out = openpyxl.Workbook()
    out.remove(out.active)
    for sheet_name in keep_sheets:
        ws = src[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        write_sheet(out, sheet_name, list(rows[0]), rows[1:])
    out.save(OUTPUTS / "Connections Data.xlsx")
    src.close()
    print("Connections Data.xlsx: " + ", ".join(keep_sheets))


if __name__ == "__main__":
    build_baidu_file()
    build_wikipedia_file()
    build_merged_file()
    build_connections_file()
