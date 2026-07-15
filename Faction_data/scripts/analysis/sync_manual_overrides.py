"""
Run this after you've hand-edited the "Manual PBSCs" column on the "Top
Connectors" sheet in Connections Data.xlsx.

It reads your edits, then recomputes and rewrites everything that depends on
manual status -- Manual Count and Score on Top Connectors, the Manual column
and Score on Top Pair Evidence, the Manual vs Bot sheet, and the Summary
counts -- using the exact same aggregation code build_unified_connections.py
uses (aggregate_and_write()), so the two can't drift apart.

It does NOT re-run the Baidu/Wikipedia matching -- it reads the complete,
uncapped evidence from the "All Candidates (raw)" sheet (written by the last
full build) and just re-applies the +50 manual bonus according to your
edited Manual PBSCs column. Fast, no network calls.

If you add a Manual PBSCs entry for a person-PBSC pair that has zero
mechanical evidence at all (never appeared in any previous build), this
creates a fresh manual-only row for it (Base Score 0, Pairs 0) -- same as
a brand-new manual_only case.

Usage: python3 sync_manual_overrides.py
"""

from pathlib import Path

import openpyxl

from build_unified_connections import OUT_XLSX, aggregate_and_write


def split_pbscs(cell_value):
    return {p.strip() for p in str(cell_value or "").split(";") if p.strip()}


def main():
    wb = openpyxl.load_workbook(OUT_XLSX, data_only=True)

    if "All Candidates (raw)" not in wb.sheetnames:
        raise SystemExit(
            "No 'All Candidates (raw)' sheet found -- run build_unified_connections.py "
            "at least once first (this file must have been produced by the current "
            "version of that script)."
        )

    raw_ws = wb["All Candidates (raw)"]
    raw_rows = list(raw_ws.iter_rows(values_only=True))
    raw_headers = list(raw_rows[0])
    idx = {h: i for i, h in enumerate(raw_headers)}

    candidate_rows = []
    title_by_name = {}
    pinyin_by_person = {}
    for r in raw_rows[1:]:
        row = {
            "Person Chinese Name": r[idx["Person Chinese Name"]],
            "Person Pinyin Name": r[idx["Person Pinyin Name"]],
            "PBSC Chinese Name": r[idx["PBSC Chinese Name"]],
            "PBSC Pinyin": r[idx["PBSC Pinyin"]],
            "Base Score": r[idx["Base Score"]] or 0,
            "Score": r[idx["Score"]] or 0,
            "Pairs": r[idx["Pairs"]] or 0,
            "Shared Anchors": r[idx["Shared Anchors"]] or "",
            "Earliest Start": r[idx["Earliest Start"]],
            "Latest End": r[idx["Latest End"]],
            "Explicit Faction Workbook Match": r[idx["Explicit Faction Workbook Match"]] or "",
            "Faction Workbook Note": r[idx["Faction Workbook Note"]] or "",
            "Evidence": r[idx["Evidence"]] or "",
        }
        candidate_rows.append(row)
        pinyin_by_person[row["Person Chinese Name"]] = row["Person Pinyin Name"]

    # Title isn't on the raw sheet -- pull it from Top Connectors, which also
    # has it, alongside the hand-edited Manual PBSCs column we actually need.
    tc_ws = wb["Top Connectors"]
    tc_rows = list(tc_ws.iter_rows(values_only=True))
    tc_headers = list(tc_rows[0])
    tc_idx = {h: i for i, h in enumerate(tc_headers)}

    manual_keys = set()
    existing_pairs = {(r["Person Chinese Name"], r["PBSC Pinyin"]) for r in candidate_rows}
    for r in tc_rows[1:]:
        person = r[tc_idx["Person"]]
        title_by_name[person] = r[tc_idx["Title"]]
        pinyin_by_person.setdefault(person, r[tc_idx["Pinyin"]])
        for pbsc_pinyin in split_pbscs(r[tc_idx["Manual PBSCs"]]):
            manual_keys.add((person, pbsc_pinyin))

    # Also pick up Title for anyone who only appears via All Candidates (raw)
    # (e.g. manual-only people who might not currently be in Top Connectors).
    src_titles_ws = None
    if "Title" not in tc_idx:
        pass  # title_by_name already best-effort populated above

    new_rows = 0
    for person, pbsc_pinyin in manual_keys:
        if (person, pbsc_pinyin) not in existing_pairs:
            pbsc_chinese = next(
                (r["PBSC Chinese Name"] for r in candidate_rows if r["PBSC Pinyin"] == pbsc_pinyin),
                pbsc_pinyin,
            )
            candidate_rows.append({
                "Person Chinese Name": person,
                "Person Pinyin Name": pinyin_by_person.get(person, ""),
                "PBSC Chinese Name": pbsc_chinese,
                "PBSC Pinyin": pbsc_pinyin,
                "Base Score": 0,
                "Score": 0,
                "Pairs": 0,
                "Shared Anchors": "",
                "Earliest Start": "",
                "Latest End": "",
                "Explicit Faction Workbook Match": "",
                "Faction Workbook Note": "manually added via Top Connectors Manual PBSCs edit",
                "Evidence": "",
            })
            existing_pairs.add((person, pbsc_pinyin))
            new_rows += 1

    # Summary's "manual_assignments" previously counted rows in the separate
    # faction workbook; post-sync it's not meaningfully different from
    # manual_candidate_rows, so just report the count of manual_keys here.
    manual_assignments_count = len(manual_keys)
    cc_people_count = wb["Summary"].cell(2, 2).value or len(candidate_rows)

    # Baidu/Wikipedia data-coverage stats (baidu_episode_rows, baidu_gap_rows,
    # wikipedia_*) don't change from a manual-only edit -- carry forward
    # whatever the last full build computed instead of losing them here,
    # since this script never reloads the source/Wikipedia data.
    summary_ws = wb["Summary"]
    carried_forward_keys = {
        "baidu_episode_rows", "baidu_gap_rows", "wikipedia_zh_matched_people",
        "wikipedia_en_matched_people", "wikipedia_episode_rows", "people_with_wikipedia_episodes_in_pool",
    }
    extra_summary = {
        row[0]: row[1] for row in summary_ws.iter_rows(min_row=2, values_only=True)
        if row[0] in carried_forward_keys
    }
    wb.close()

    if new_rows:
        print(f"Added {new_rows} new manual-only candidate row(s) not previously in All Candidates (raw).")
    print(f"Applying {len(manual_keys)} manual person-PBSC pair(s) from Top Connectors' Manual PBSCs column...")

    aggregate_and_write(candidate_rows, manual_keys, title_by_name, cc_people_count, manual_assignments_count, OUT_XLSX, extra_summary)


if __name__ == "__main__":
    main()
