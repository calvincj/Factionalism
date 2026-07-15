"""
Rebuilds Connections Data.xlsx from BOTH Baidu and Wikipedia career data,
pooled together as one set of evidence per person rather than scored
separately.

For each person, builds one combined list of "episodes" (anchor + year
interval), one source being their Baidu career episodes, the other being
their Wikipedia infobox offices (parsed fresh from the cached API responses,
using the person/page match already resolved by enrich_wikipedia.py). Anchor
detection reuses enrich_baidu.py's Chinese place/institution lists directly
for Chinese-language episodes and offices; English Wikipedia offices are
matched against a parallel English->anchor table.

Then it runs the exact same pairwise anchor+interval-overlap matching and
scoring formula enrich_baidu.py uses for Baidu-only data, just against the
pooled episodes -- so a "pair" can now be Baidu-vs-Baidu, Baidu-vs-Wikipedia,
or Wikipedia-vs-Wikipedia. No column distinguishes which source contributed
a given pair or score; the per-pair evidence text keeps a short [Baidu] /
[Wikipedia-ZH] / [Wikipedia-EN] tag per line so provenance is still
traceable if you want to check a specific claim.

Score is split into "Base Score" (evidence-derived: pairs + anchors +
systems, independent of manual status) and the +50 manual-workbook bonus,
which is applied last. This split is what lets sync_manual_overrides.py
recompute everything after a hand-edit to Top Connectors' "Manual PBSCs"
column without re-deriving evidence from scratch: it only needs to add/
remove the 50-point bonus per pair, never touch Base Score.

The aggregation/sheet-writing logic lives in aggregate_and_write() so both
this script and sync_manual_overrides.py call the exact same code -- there
is no separate, divergence-prone copy of the scoring/rollup logic.
"""

import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR.parent / "data_collection"))
import enrich_baidu as eb  # noqa: E402
from enrich_wikipedia import api_get, extract_page, parse_infobox_offices  # noqa: E402

FACTION_DATA_DIR = SCRIPT_DIR.parent.parent  # scripts/analysis/ -> scripts/ -> Faction_data/
SOURCE_WORKBOOK = FACTION_DATA_DIR / "source" / "Chinese Leadership Database - Baidu enriched working copy.xlsx"
FACTION_WORKBOOK = FACTION_DATA_DIR / "source" / "PBSC Factionalism - working copy.xlsx"
WIKI_DATA_XLSX = FACTION_DATA_DIR / "outputs" / "Wikipedia Data - All People.xlsx"
OUT_XLSX = FACTION_DATA_DIR / "outputs" / "Connections Data.xlsx"

# Point enrich_baidu's module-level constant at our real local file --
# parse_pbsc_faction_workbook() reads this at call time, so patching the
# attribute before calling it is enough; no need to touch enrich_baidu.py.
eb.FACTION_WORKBOOK = FACTION_WORKBOOK

EN_PROVINCE_MAP = {
    "Beijing": "北京", "Tianjin": "天津", "Shanghai": "上海", "Chongqing": "重庆",
    "Hebei": "河北", "Shanxi": "山西", "Liaoning": "辽宁", "Jilin": "吉林",
    "Heilongjiang": "黑龙江", "Jiangsu": "江苏", "Zhejiang": "浙江", "Anhui": "安徽",
    "Fujian": "福建", "Jiangxi": "江西", "Shandong": "山东", "Henan": "河南",
    "Hubei": "湖北", "Hunan": "湖南", "Guangdong": "广东", "Hainan": "海南",
    "Sichuan": "四川", "Guizhou": "贵州", "Yunnan": "云南", "Shaanxi": "陕西",
    "Gansu": "甘肃", "Qinghai": "青海", "Taiwan": "台湾", "Inner Mongolia": "内蒙古",
    "Guangxi": "广西", "Tibet": "西藏", "Ningxia": "宁夏", "Xinjiang": "新疆",
    "Hong Kong": "香港", "Macau": "澳门", "Macao": "澳门",
}

EN_INSTITUTION_MAP = {
    "Tsinghua University": "清华大学", "Peking University": "北京大学",
    "Fudan University": "复旦大学", "Zhejiang University": "浙江大学",
    "Nankai University": "南开大学", "Jilin University": "吉林大学",
    "Wuhan University": "武汉大学", "Nanjing University": "南京大学",
    "Xiamen University": "厦门大学", "Renmin University": "中国人民大学",
    "Chinese Academy of Social Sciences": "中国社会科学院",
    "Chinese Academy of Sciences": "中国科学院",
    "Chinese Academy of Engineering": "中国工程院",
    "Central Party School": "中央党校", "Organization Department": "中央组织部",
    "Publicity Department": "中央宣传部", "United Front Work Department": "中央统战部",
    "Ministry of Foreign Affairs": "外交部", "Ministry of Finance": "财政部",
    "Ministry of Commerce": "商务部", "Ministry of Public Security": "公安部",
    "Ministry of Justice": "司法部", "General Administration of Customs": "海关总署",
    "National Development and Reform Commission": "国家发展改革委",
    "China Development Bank": "国家开发银行",
}
EN_ANCHOR_MAP = {**EN_PROVINCE_MAP, **EN_INSTITUTION_MAP}


def detect_places_en(text):
    text = text or ""
    hits = []
    for en, cn in EN_ANCHOR_MAP.items():
        if re.search(rf"\b{re.escape(en)}\b", text, re.IGNORECASE) and cn not in hits:
            hits.append(cn)
    return hits


def parse_wiki_year(text):
    text = str(text or "")
    if re.search(r"present|incumbent|至今|现任|Incumbent", text, re.IGNORECASE):
        return None
    m = re.search(r"(\d{4})", text)
    return int(m.group(1)) if m else None


def rows_as_dicts(ws):
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(x or "").strip() for x in rows[0]]
    return [{headers[i]: v[i] if i < len(v) else None for i in range(len(headers))} for v in rows[1:]]


def load_baidu_episodes_by_name():
    wb = openpyxl.load_workbook(SOURCE_WORKBOOK, read_only=True, data_only=True)
    rows = rows_as_dicts(wb["Baidu Career Episodes"])
    by_name = defaultdict(list)
    for r in rows:
        name = r.get("Chinese Name")
        if not name:
            continue
        by_name[str(name).strip()].append(r)
    wb.close()
    return by_name


def load_wiki_titles_by_name():
    wb = openpyxl.load_workbook(WIKI_DATA_XLSX, read_only=True, data_only=True)
    rows = rows_as_dicts(wb["People"])
    titles = {}
    for r in rows:
        name = r.get("Chinese Name")
        if not name:
            continue
        name = str(name).strip()
        titles[(name, "zh")] = (r.get("Wikipedia ZH Status"), r.get("Wikipedia ZH Title"))
        titles[(name, "en")] = (r.get("Wikipedia EN Status"), r.get("Wikipedia EN Title"))
    wb.close()
    return titles


GOOD_WIKI_STATUSES = {"matched", "matched_via_disambiguation"}


_normalized_index_cache = {}


def _normalized_lookup_index(by_name_dict, key_fn=lambda k: k):
    """Maps normalize_name(key) -> the dict's value, but ONLY when exactly
    one raw key produces that normalized form. Two distinct people can share
    a normalized name (e.g. two different officials both literally named
    "王凯", disambiguated only by "(河南)" vs "（解放军）") -- for those, skip
    the fallback entirely rather than risk silently merging two different
    people's career data. Cached per dict identity since this gets called
    once per person (~92+ times) against the same two dicts.
    """
    cache_key = id(by_name_dict)
    if cache_key not in _normalized_index_cache:
        groups = defaultdict(list)
        for k in by_name_dict:
            groups[eb.normalize_name(key_fn(k))].append(k)
        _normalized_index_cache[cache_key] = {
            norm: by_name_dict[keys[0]] for norm, keys in groups.items() if len(keys) == 1
        }
    return _normalized_index_cache[cache_key]


def build_episode_pool(chinese_name, baidu_by_name, wiki_titles):
    """chinese_name may or may not carry a gender/ethnicity parenthetical
    (e.g. "程丽华（女）") depending on which sheet it came from -- Baidu Career
    Episodes keeps these, but enrich_wikipedia.py strips them when building
    Wikipedia Data - All People.xlsx. Without normalizing, a name with a
    parenthetical (17 people, confirmed) would find its Baidu episodes but
    silently get zero Wikipedia episodes even when good Wikipedia coverage
    exists (or vice versa), because the two dicts disagree on the person's
    key. Try the exact name first, then an unambiguous normalized-name
    fallback (see _normalized_lookup_index).
    """
    episodes = []

    baidu_rows = baidu_by_name.get(chinese_name)
    if baidu_rows is None:
        baidu_rows = _normalized_lookup_index(baidu_by_name).get(eb.normalize_name(chinese_name), [])
    for r in baidu_rows:
        interval = eb.year_interval(r.get("Start Year"), r.get("End Year"))
        if not interval:
            continue
        places = eb.split_tags(r.get("Detected Places"))
        anchors = []
        for anchor in places:
            if anchor in eb.EXCLUDED_OVERLAP_ANCHORS:
                continue
            anchors.append({"anchor": anchor, "interval": eb.anchor_interval(anchor, r.get("Raw Timeline Entry (CN)"), interval)})
        if not anchors:
            continue
        episodes.append({
            "anchors": anchors,
            "interval": interval,
            "systems": eb.split_tags(r.get("Detected System Tags")),
            "source": "Baidu",
            "text": f"{r.get('Start Year')}-{r.get('End Year') or ''} {r.get('Role Text (CN)') or ''}".strip(),
        })

    for lang in ("zh", "en"):
        entry = wiki_titles.get((chinese_name, lang))
        if entry is None:
            # wiki_titles always has both a zh and en tuple-key per person, so
            # every name is technically "ambiguous" once grouped by normalized
            # name alone -- filter to this lang first, then require exactly
            # one distinct normalized name within it before trusting the match.
            norm = eb.normalize_name(chinese_name)
            candidates = {n for (n, l) in wiki_titles if l == lang and eb.normalize_name(n) == norm}
            entry = wiki_titles.get((next(iter(candidates)), lang)) if len(candidates) == 1 else None
        status, title = entry or (None, None)
        if status not in GOOD_WIKI_STATUSES or not title:
            continue
        data = api_get(lang, title)
        page = extract_page(data)
        if not page:
            continue
        for o in parse_infobox_offices(page["wikitext"]):
            start = parse_wiki_year(o["term_start"])
            if start is None:
                continue
            end = parse_wiki_year(o["term_end"])
            interval = eb.year_interval(start, end)
            if not interval:
                continue
            detect_fn = eb.detect_places if lang == "zh" else detect_places_en
            raw_anchors = eb.split_tags(detect_fn(o["office"])) if lang == "zh" else detect_fn(o["office"])
            anchors = [{"anchor": a, "interval": interval} for a in raw_anchors if a not in eb.EXCLUDED_OVERLAP_ANCHORS]
            if not anchors:
                continue
            tag = f"Wikipedia-{lang.upper()}"
            episodes.append({
                "anchors": anchors,
                "interval": interval,
                "systems": set(),
                "source": tag,
                "text": f"{o['term_start']}-{o['term_end'] or 'present'} {o['office']}".strip(),
            })

    return episodes


def compute_candidate_rows(all_names, baidu_by_name, wiki_titles):
    """Runs the pooled Baidu+Wikipedia anchor/interval matching. Score in the
    returned rows does NOT include the manual +50 bonus -- that is Base Score.
    Manual flag/note here just reflects whatever's currently in
    source/PBSC Factionalism - working copy.xlsx at call time; callers that
    want a different manual set (e.g. sync_manual_overrides.py) should ignore
    Explicit Faction Workbook Match / Faction Workbook Note and recompute
    Score from Base Score against their own manual_keys.
    """
    assignments_by_name = eb.parse_pbsc_faction_workbook()
    pinyin_by_name = {n: p for n, p in all_names}

    print(f"Building combined episode pools for {len(all_names)} people...")
    pools = {name: build_episode_pool(name, baidu_by_name, wiki_titles) for name, _ in all_names}

    pbsc_pools = {name: pools[name] for name in eb.KNOWN_PBSC_CN if name in pools}
    people = [(n, p) for n, p in all_names if n not in eb.KNOWN_PBSC_CN]

    candidate_rows = []
    for person_name, person_pinyin in people:
        person_eps = pools.get(person_name, [])
        # No early skip on empty person_eps: someone with zero qualifying
        # episodes (thin/purged-official coverage) can still have a
        # manual-only faction-workbook match, which is checked per-PBSC below.
        for pbsc_name in sorted(eb.KNOWN_PBSC_CN):
            pbsc_eps = pbsc_pools.get(pbsc_name, [])
            if not pbsc_eps:
                continue
            faction_hits = [
                item for item in assignments_by_name.get(eb.normalize_name(person_name), [])
                if item["pbsc_chinese"] == pbsc_name
            ]
            faction_hit_note = eb.format_faction_matches(faction_hits) if faction_hits else ""

            shared_anchors, shared_systems, evidence, overlap_windows = set(), set(), [], []
            pair_count = 0
            for left in person_eps:
                for right in pbsc_eps:
                    if not eb.intervals_overlap(left["interval"], right["interval"]):
                        continue
                    systems = left["systems"] & right["systems"]
                    matches = []
                    for la in left["anchors"]:
                        for ra in right["anchors"]:
                            if la["anchor"] != ra["anchor"]:
                                continue
                            if not eb.intervals_overlap(la["interval"], ra["interval"]):
                                continue
                            matches.append((la["anchor"], eb.overlap_window(la["interval"], ra["interval"])))
                    if not matches:
                        continue
                    pair_count += 1
                    shared_systems.update(systems)
                    for anchor, window in matches:
                        shared_anchors.add(anchor)
                        overlap_windows.append(window)
                    if len(evidence) < 5:
                        anchor_text = ", ".join(sorted({f"{a} {w[0]}-{w[1]}" for a, w in matches}))
                        evidence.append(
                            f"{anchor_text}: [{left['source']}] {left['text']} ↔ [{right['source']}] {pbsc_name} {right['text']}"
                        )
            if pair_count or faction_hits:
                base_score = (5 * pair_count) + (10 * len(shared_anchors)) + (3 * len(shared_systems))
                candidate_rows.append({
                    "Person Chinese Name": person_name,
                    "Person Pinyin Name": person_pinyin,
                    "PBSC Chinese Name": pbsc_name,
                    "PBSC Pinyin": pinyin_by_name.get(pbsc_name, pbsc_name),
                    "Explicit Faction Workbook Match": "Y" if faction_hits else "",
                    "Faction Workbook Note": faction_hit_note,
                    "Base Score": base_score,
                    "Score": base_score + (50 if faction_hits else 0),
                    "Pairs": pair_count,
                    "Shared Anchors": "; ".join(sorted(shared_anchors)),
                    "Earliest Start": min((s for s, _ in overlap_windows), default=""),
                    "Latest End": max((e for _, e in overlap_windows), default=""),
                    "Evidence": "\n".join(evidence),
                })

    print(f"Candidate rows: {len(candidate_rows)}")
    return candidate_rows, pools


def anchor_type(anchor):
    return "province/place" if anchor in eb.PROVINCE_TERMS else "institution/org"


def autosize(ws, max_width=90):
    for col in ws.columns:
        letter = col[0].column_letter
        width = max([10] + [min(max_width, len(str(c.value or "")) + 2) for c in col[:200]])
        ws.column_dimensions[letter].width = width


def write_sheet(wb, name, headers, rows):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    autosize(ws)
    return ws


def aggregate_and_write(candidate_rows, manual_keys, title_by_name, cc_people_count, manual_assignments_count, out_path, extra_summary=None):
    """Recomputes Score from Base Score against manual_keys, then rebuilds
    every sheet. manual_keys is the single source of truth for "is this
    person-PBSC pair manual" -- both the full rebuild (main(), manual_keys
    from the faction workbook) and sync_manual_overrides.py (manual_keys from
    a hand-edited Top Connectors sheet) go through this same function so the
    two code paths can't drift apart.
    """
    for r in candidate_rows:
        key = (r["Person Chinese Name"], r["PBSC Pinyin"])
        is_manual = key in manual_keys
        r["Explicit Faction Workbook Match"] = "Y" if is_manual else ""
        r["Score"] = r["Base Score"] + (50 if is_manual else 0)

    physical_rows = [r for r in candidate_rows if r["Shared Anchors"]]
    manual_rows = [r for r in candidate_rows if r["Explicit Faction Workbook Match"] == "Y"]
    manual_no_physical = [r for r in manual_rows if not r["Shared Anchors"]]

    by_pbsc = {}
    for pbsc in sorted(eb.KNOWN_PBSC_CN):
        rows = [r for r in candidate_rows if r["PBSC Chinese Name"] == pbsc]
        phys = [r for r in rows if r["Shared Anchors"]]
        manual = [r for r in rows if r["Explicit Faction Workbook Match"] == "Y"]
        places, institutions = Counter(), Counter()
        for r in phys:
            for a in r["Shared Anchors"].split("; "):
                if not a:
                    continue
                (places if anchor_type(a) == "province/place" else institutions)[a] += 1
        by_pbsc[pbsc] = {
            "people": len({r["Person Chinese Name"] for r in phys}),
            "manual_people": len({r["Person Chinese Name"] for r in manual}),
            "pairs": sum(r["Pairs"] for r in phys),
            "top_places": places.most_common(6),
            "top_institutions": institutions.most_common(6),
        }

    by_person = defaultdict(lambda: {"pinyin": "", "pbscs": set(), "manual_pbscs": set(), "score": 0, "pairs": 0, "anchors": Counter()})
    for r in candidate_rows:
        rec = by_person[r["Person Chinese Name"]]
        rec["pinyin"] = r["Person Pinyin Name"]
        if r["Shared Anchors"]:
            rec["pbscs"].add(r["PBSC Pinyin"])
            rec["score"] += r["Score"]
            rec["pairs"] += r["Pairs"]
            for a in r["Shared Anchors"].split("; "):
                if a:
                    rec["anchors"][a] += 1
        if r["Explicit Faction Workbook Match"] == "Y":
            rec["manual_pbscs"].add(r["PBSC Pinyin"])

    top_connectors = sorted(
        [
            {
                "person": p, "pinyin": rec["pinyin"], "title": title_by_name.get(p, ""),
                "pbsc_count": len(rec["pbscs"]), "manual_count": len(rec["manual_pbscs"]),
                "score": rec["score"], "pairs": rec["pairs"],
                "anchors": "; ".join(a for a, _ in rec["anchors"].most_common(8)),
                "pbscs": "; ".join(sorted(rec["pbscs"])), "manual_pbscs": "; ".join(sorted(rec["manual_pbscs"])),
            }
            for p, rec in by_person.items() if rec["pbscs"] or rec["manual_pbscs"]
        ],
        key=lambda x: (x["pbsc_count"], x["score"], x["pairs"]), reverse=True,
    )

    top_pairs = sorted(physical_rows, key=lambda r: (r["Score"], r["Pairs"]), reverse=True)

    anchor_counts, anchor_pair_counts = Counter(), Counter()
    for r in physical_rows:
        for a in r["Shared Anchors"].split("; "):
            if a:
                anchor_counts[a] += 1
                anchor_pair_counts[a] += r["Pairs"]

    physical_keys = {(r["Person Chinese Name"], r["PBSC Pinyin"]) for r in physical_rows}

    summary = {
        "central_committee_rows": cc_people_count,
        "candidate_rows_total": len(candidate_rows),
        "physical_candidate_rows": len(physical_rows),
        "manual_candidate_rows": len(manual_rows),
        "manual_without_physical_rows": len(manual_no_physical),
        "unique_people_with_overlap": len({r["Person Chinese Name"] for r in physical_rows}),
        "manual_assignments": manual_assignments_count,
    }
    if extra_summary:
        summary.update(extra_summary)

    out = openpyxl.Workbook()
    out.remove(out.active)
    write_sheet(out, "Summary", ["Metric", "Value"], [[k, v] for k, v in summary.items()])
    write_sheet(out, "By PBSC", ["PBSC", "People_bot", "People_manual", "Pairs", "Top Places", "Top Institutions"], [
        [pbsc, info["people"], info["manual_people"], info["pairs"],
         "; ".join(f"{a} ({c})" for a, c in info["top_places"]),
         "; ".join(f"{a} ({c})" for a, c in info["top_institutions"])]
        for pbsc, info in by_pbsc.items()
    ])
    write_sheet(out, "Top Connectors", ["Person", "Pinyin", "Title", "PBSC Count", "Manual Count", "Score", "Pairs", "PBSCs", "Manual PBSCs", "Top Anchors"], [
        [r["person"], r["pinyin"], r["title"], r["pbsc_count"], r["manual_count"], r["score"], r["pairs"], r["pbscs"], r["manual_pbscs"], r["anchors"]]
        for r in top_connectors
    ])
    write_sheet(out, "Top Pair Evidence", ["Person", "Pinyin", "Title", "PBSC", "Manual", "Score", "Pairs", "Anchors", "Start", "End", "Evidence"], [
        [r["Person Chinese Name"], r["Person Pinyin Name"], title_by_name.get(r["Person Chinese Name"], ""), r["PBSC Pinyin"],
         r["Explicit Faction Workbook Match"], r["Score"], r["Pairs"], r["Shared Anchors"], r["Earliest Start"], r["Latest End"], r["Evidence"]]
        for r in top_pairs
    ])
    write_sheet(out, "Anchors", ["Anchor", "Type", "Candidate Rows", "Pair Count"], [
        [a, anchor_type(a), c, anchor_pair_counts[a]] for a, c in anchor_counts.most_common()
    ])
    person_en_by_cn = {r["Person Chinese Name"]: r["Person Pinyin Name"] for r in candidate_rows}
    pbsc_cn_by_en = {r["PBSC Pinyin"]: r["PBSC Chinese Name"] for r in candidate_rows}
    note_by_pair = {(r["Person Chinese Name"], r["PBSC Pinyin"]): r["Faction Workbook Note"] for r in candidate_rows}

    def mvb_row(found_by, person_cn, pbsc_en):
        return [
            found_by, person_cn, pbsc_cn_by_en.get(pbsc_en, pbsc_en), person_en_by_cn.get(person_cn, ""), pbsc_en,
            note_by_pair.get((person_cn, pbsc_en), ""),
        ]

    write_sheet(out, "Manual vs Bot", ["Found By", "Person_cn", "PBSC_cn", "Person_en", "PBSC_en", "Notes"],
                [mvb_row("manual_and_bot", p, pb) for p, pb in sorted(manual_keys & physical_keys)]
                + [mvb_row("manual", p, pb) for p, pb in sorted(manual_keys - physical_keys)]
                + [mvb_row("bot", p, pb) for p, pb in sorted(physical_keys - manual_keys)])

    # Complete candidate data, same rows as Top Pair Evidence above (both are
    # now uncapped) but including manual-only rows with no physical overlap
    # too -- sync_manual_overrides.py reads from here.
    write_sheet(out, "All Candidates (raw)", [
        "Person Chinese Name", "Person Pinyin Name", "PBSC Chinese Name", "PBSC Pinyin",
        "Base Score", "Score", "Pairs", "Shared Anchors", "Earliest Start", "Latest End",
        "Explicit Faction Workbook Match", "Faction Workbook Note", "Evidence",
    ], [
        [r["Person Chinese Name"], r["Person Pinyin Name"], r["PBSC Chinese Name"], r["PBSC Pinyin"],
         r["Base Score"], r["Score"], r["Pairs"], r["Shared Anchors"], r["Earliest Start"], r["Latest End"],
         r["Explicit Faction Workbook Match"], r["Faction Workbook Note"], r["Evidence"]]
        for r in candidate_rows
    ])

    write_sheet(out, "Data Dictionary", ["Field", "Definition"], [
        ["Methodology", "Baidu career episodes and Wikipedia (ZH+EN) infobox offices are pooled into one combined set of career evidence per person before matching -- a 'pair' can come from either source, and no column distinguishes which. See Evidence text for per-line [Baidu]/[Wikipedia-ZH]/[Wikipedia-EN] source tags."],
        ["Anchor", "A specific shared place (province/municipality) or named institution (university, central party organ, ministry, etc.) that appears in both the person's and the PBSC member's career records, in overlapping years."],
        ["Pairs", "Count of matched (person episode) x (PBSC member episode) combinations sharing an anchor with overlapping years. NOT the same as years overlapped or distinct anchors shared -- a long stretch chopped into many source entries on either side multiplies into more pairs even if it reflects one real continuous overlap. Treat as a rough volume-of-evidence signal, not a precision measure."],
        ["Anchors (count/list)", "Number/list of DISTINCT shared anchors (unlike Pairs, each anchor only counts once regardless of how many episode-pairs matched on it). More informative than Pairs for judging breadth of connection."],
        ["Top Places / Top Institutions (By PBSC sheet)", "The number in parentheses is a count of how many DIFFERENT connected people share that anchor with this PBSC member (i.e. breadth -- capped at People_bot) -- NOT a count of matched episode-pairs. It will not add up to the Pairs column, which sums cross-multiplied episode-matches across all anchors and all connected people. Top Places = province/municipality anchors; Top Institutions = named universities/central organs/ministries."],
        ["People_bot / People_manual (By PBSC sheet)", "People_bot = number of distinct people mechanically matched to this PBSC member (Shared Anchors non-empty). People_manual = number of distinct people hand-assigned to this PBSC member in the faction workbook, regardless of whether a mechanical match was also found. The two are not mutually exclusive subsets of each other -- see Manual vs Bot for the person-level overlap."],
        ["Base Score", "5 x Pairs + 10 x (distinct shared anchors) + 3 x (shared broad system tags, Baidu episodes only). Evidence-derived only -- does NOT include the manual bonus."],
        ["Score", "Base Score + 50 if the pair is manually flagged (either in the hand-coded PBSC Factionalism workbook, or via edits to Top Connectors' Manual PBSCs column after running sync_manual_overrides.py). An arbitrary hand-tuned weighted index, not a statistically validated measure -- see conversation notes on manual-bonus circularity and lack of a chance/null-model baseline."],
        ["Manual / Explicit Faction Workbook Match", "Y if this person-PBSC pair is currently flagged manual, independent of anything mechanically detected here."],
        ["Evidence", "Up to 5 sample matched episode-pairs (anchor, overlap window, and both people's role text at the time), each tagged with its source. Not exhaustive -- there may be more matches than shown."],
        ["Manual vs Bot 'Found By' values", "manual_and_bot = hand-assigned AND mechanically matched (highest confidence, two independent methods agree); manual = hand-assigned but no mechanical match found; bot = mechanically matched but not currently flagged manual. Uncapped -- every pair is listed."],
        ["All Candidates (raw)", "Every candidate row, including manual-only rows with no physical overlap that Top Pair Evidence excludes (Top Pair Evidence is otherwise the same uncapped set). This is what sync_manual_overrides.py reads to recompute everything after you hand-edit Manual PBSCs in Top Connectors, without re-running the Baidu/Wikipedia matching."],
        ["Summary: baidu_episode_rows / baidu_gap_rows", "Raw counts from the source data, unrelated to matching: total parsed Baidu career-episode rows across all 219 people, and how many still need non-Baidu fallback research (Baidu Gap Review sheet in the source workbook)."],
        ["Summary: wikipedia_zh_matched_people / wikipedia_en_matched_people", "How many of the 219 people were successfully resolved to a Chinese / English Wikipedia page (status matched or matched_via_disambiguation) -- coverage, not evidence quality."],
        ["Summary: wikipedia_episode_rows / people_with_wikipedia_episodes_in_pool", "wikipedia_episode_rows = total Wikipedia-sourced office entries that made it into ANY combined episode pool (had a parseable date and a detected anchor). people_with_wikipedia_episodes_in_pool = how many distinct people had at least one such usable entry -- smaller than wikipedia_zh/en_matched_people because many matched Wikipedia pages have offices with no detectable place/institution anchor or no parseable date."],
        ["Caveat", "A shared anchor + overlapping years is evidence of possible career contact, not proof of a factional tie. Populous anchors (large provinces) generate many pairs largely by chance given how many officials rotate through them -- see conversation notes on the homophily/base-rate problem before treating Score as a ranked measure of relationship strength."],
    ])

    out.save(out_path)
    print(f"Wrote {out_path}")
    print(summary)


def main():
    src = openpyxl.load_workbook(SOURCE_WORKBOOK, read_only=True, data_only=True)
    cc_rows = rows_as_dicts(src["20th Central Committee"])
    # Source sheet has at least one name with stray trailing whitespace (Wang
    # Huning's Chinese Name is "王沪宁 ") which silently breaks dict-key lookups
    # against KNOWN_PBSC_CN/Baidu-episode names elsewhere -- strip everywhere.
    title_by_name = {str(r.get("Chinese Name")).strip(): r.get("Title") for r in cc_rows if r.get("Chinese Name")}
    all_names = [(str(r.get("Chinese Name")).strip(), str(r.get("Pinyin Name") or "").strip()) for r in cc_rows if r.get("Chinese Name")]
    src.close()

    baidu_by_name = load_baidu_episodes_by_name()
    wiki_titles = load_wiki_titles_by_name()

    candidate_rows, pools = compute_candidate_rows(all_names, baidu_by_name, wiki_titles)

    assignments_by_name = eb.parse_pbsc_faction_workbook()
    manual_assignments_count = sum(len(v) for v in assignments_by_name.values())
    manual_keys = {
        (r["Person Chinese Name"], r["PBSC Pinyin"])
        for r in candidate_rows
        if r["Explicit Faction Workbook Match"] == "Y"
    }

    # Baidu Gap Review row count -- restored from the original Baidu-only summary.
    src2 = openpyxl.load_workbook(SOURCE_WORKBOOK, read_only=True, data_only=True)
    baidu_gap_rows = max(0, src2["Baidu Gap Review"].max_row - 1)
    src2.close()

    wiki_zh_matched = sum(1 for n, _ in all_names if wiki_titles.get((n, "zh"), (None,))[0] in GOOD_WIKI_STATUSES)
    wiki_en_matched = sum(1 for n, _ in all_names if wiki_titles.get((n, "en"), (None,))[0] in GOOD_WIKI_STATUSES)
    wiki_episode_rows = sum(1 for eps in pools.values() for ep in eps if ep["source"].startswith("Wikipedia"))
    people_with_wiki_episodes = sum(1 for eps in pools.values() if any(ep["source"].startswith("Wikipedia") for ep in eps))

    extra_summary = {
        "baidu_episode_rows": sum(len(v) for v in baidu_by_name.values()),
        "baidu_gap_rows": baidu_gap_rows,
        "wikipedia_zh_matched_people": wiki_zh_matched,
        "wikipedia_en_matched_people": wiki_en_matched,
        "wikipedia_episode_rows": wiki_episode_rows,
        "people_with_wikipedia_episodes_in_pool": people_with_wiki_episodes,
    }

    aggregate_and_write(candidate_rows, manual_keys, title_by_name, len(all_names), manual_assignments_count, OUT_XLSX, extra_summary)


if __name__ == "__main__":
    main()
