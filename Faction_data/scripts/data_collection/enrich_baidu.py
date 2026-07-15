import html
import hashlib
import json
import re
import subprocess
import time
import urllib.parse
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from bs4 import BeautifulSoup


ROOT = Path(r"C:\Users\AMatthias\Documents\Codex\2026-05-12\i-want-to-make-database-that")
WORKBOOK = ROOT / "Chinese Leadership Database - Baidu enriched working copy.xlsx"
FACTION_WORKBOOK = ROOT / "PBSC Factionalism - working copy.xlsx"
CACHE_DIR = ROOT / "baidu_cache"
REPORT = ROOT / "baidu_enrichment_report.json"

MAIN_SHEET = "20th Central Committee"
EPISODE_SHEET = "Baidu Career Episodes"
PBSC_OVERLAP_SHEET = "PBSC Overlap Prep"
PBSC_PAIR_SHEET = "PBSC Overlap Candidates"
GAP_SHEET = "Baidu Gap Review"
FACTION_SHEET = "PBSC Faction Assignments"
CURRENT_YEAR = 2026

ADDED_COLUMNS = [
    "Baidu Query Name",
    "Baidu URL",
    "Baidu Last Updated",
    "Baidu Profile Summary (CN)",
    "Baidu Career Timeline (CN)",
    "Baidu Timeline Episode Count",
    "Baidu Match Status",
    "Baidu Notes",
]

FACTION_COLUMNS = [
    "PBSC Faction Workbook Count",
    "PBSC Faction Workbook Matches",
    "PBSC Faction Workbook Notes",
]

PBSC_PINYIN_TO_CN = {
    "Xi Jinping": "习近平",
    "Li Qiang": "李强",
    "Zhao Leji": "赵乐际",
    "Wang Huning": "王沪宁",
    "Cai Qi": "蔡奇",
    "Ding Xuexiang": "丁薛祥",
    "Li Xi": "李希",
}

KNOWN_PBSC_CN = {
    "习近平",
    "李强",
    "赵乐际",
    "王沪宁",
    "蔡奇",
    "丁薛祥",
    "李希",
}

SYSTEM_KEYWORDS = {
    "party_center": ["中央办公厅", "中央组织部", "中央宣传部", "中央统战部", "中央政法委", "中央纪委", "中央书记处", "中央党校", "中央政策研究室", "中央和国家机关"],
    "state_council": ["国务院", "国家发展改革委", "财政部", "外交部", "商务部", "公安部", "民政部", "司法部", "审计署", "海关总署", "人民银行", "国家统计局"],
    "provincial": ["省委", "省政府", "自治区党委", "自治区政府", "市委", "市政府", "区委", "县委", "地委"],
    "pla": ["军委", "解放军", "战区", "军区", "陆军", "海军", "空军", "火箭军", "武警", "国防大学", "军事"],
    "soe": ["集团", "公司", "中石油", "中石化", "航天", "兵器", "电子科技", "国有"],
    "academia": ["大学", "学院", "研究所", "科学院", "工程院", "教授", "院士"],
    "cppcc_npc": ["全国人大", "全国政协", "人大常委会", "政协"],
}

PROVINCE_TERMS = [
    "北京", "天津", "上海", "重庆", "河北", "山西", "辽宁", "吉林", "黑龙江", "江苏", "浙江", "安徽",
    "福建", "江西", "山东", "河南", "湖北", "湖南", "广东", "海南", "四川", "贵州", "云南", "陕西",
    "甘肃", "青海", "台湾", "内蒙古", "广西", "西藏", "宁夏", "新疆", "香港", "澳门",
]

INSTITUTION_TERMS = [
    "中央党校",
    "中共中央党校",
    "国家行政学院",
    "中央社会主义学院",
    "中国浦东干部学院",
    "中国井冈山干部学院",
    "中国延安干部学院",
    "中国人民大学",
    "北京大学",
    "清华大学",
    "复旦大学",
    "浙江大学",
    "南开大学",
    "吉林大学",
    "武汉大学",
    "南京大学",
    "厦门大学",
    "哈尔滨工业大学",
    "中央财经大学",
    "中国社会科学院",
    "中国科学院",
    "中国工程院",
    "中央办公厅",
    "中央组织部",
    "中央宣传部",
    "中央统战部",
    "中央政法委",
    "中央纪委",
    "中央政策研究室",
    "外交部",
    "财政部",
    "商务部",
    "公安部",
    "司法部",
    "审计署",
    "海关总署",
    "国家发展改革委",
    "中国人民银行",
    "中国银行",
    "国家开发银行",
]

PLACE_NAME_INSTITUTION_SUFFIXES = (
    "大学",
    "学院",
    "师范",
    "理工",
    "交通",
    "财经",
    "农业",
    "工业",
    "科技",
    "外国语",
    "政法",
    "医科",
)

EXCLUDED_OVERLAP_ANCHORS = {
    "中央党校",
    "中共中央党校",
}


@dataclass
class ParsedPage:
    url: str
    canonical_url: str
    title: str
    date_updated: str
    summary: str
    timeline: list[str]
    match_status: str
    notes: str


def run_powershell_fetch(url: str) -> str:
    command = (
        "[Console]::OutputEncoding=[System.Text.Encoding]::UTF8; "
        "$ProgressPreference='SilentlyContinue'; "
        f"$r=Invoke-WebRequest -Uri '{url}' -UseBasicParsing -TimeoutSec 30; "
        "[Console]::Out.Write($r.Content)"
    )
    completed = subprocess.run(
        ["powershell", "-NoProfile", "-Command", command],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=45,
    )
    if completed.returncode != 0:
        raise RuntimeError(completed.stderr.strip() or f"PowerShell fetch failed: {completed.returncode}")
    return completed.stdout


def cache_path_for(cache_key: str) -> Path:
    CACHE_DIR.mkdir(exist_ok=True)
    slug = re.sub(r"[^\w\u4e00-\u9fff-]+", "_", cache_key).strip("_")
    if len(slug) > 80:
        slug = slug[:80]
    return CACHE_DIR / f"{slug}.html"


def fetch_baidu_url(url: str, cache_key: str) -> tuple[str, str]:
    cache_path = cache_path_for(cache_key)
    if cache_path.exists() and cache_path.stat().st_size > 5000:
        return cache_path.read_text(encoding="utf-8", errors="replace"), "cache"

    html_text = run_powershell_fetch(url)
    cache_path.write_text(html_text, encoding="utf-8")
    time.sleep(0.45)
    return html_text, "fetched"


def direct_baidu_url(chinese_name: str) -> str:
    return "https://baike.baidu.com/item/" + urllib.parse.quote(chinese_name)


def baidu_query_name(raw_name: str) -> str:
    name = (raw_name or "").strip()
    name = re.sub(r"[\(（].*?[\)）]", "", name).strip()
    name = re.sub(r"\s+", "", name)
    return name


def normalize_name(raw_name: str) -> str:
    return baidu_query_name(raw_name)


def contains_cjk(value: str) -> bool:
    return bool(re.search(r"[\u4e00-\u9fff]", str(value or "")))


def baidu_urls_from_row(ws, row: int, ignored_cols: set[int]) -> list[str]:
    urls = []
    for col in range(1, ws.max_column + 1):
        if col in ignored_cols:
            continue
        value = str(ws.cell(row, col).value or "")
        for match in re.findall(r"https?://(?:wap)?baike\.baidu\.com/item/[^\s,;，；]+", value):
            cleaned = match.rstrip(").,;，；、")
            if cleaned not in urls:
                urls.append(cleaned)
    return urls


def candidate_baidu_urls(query_name: str, source_urls: list[str]) -> list[tuple[str, str]]:
    candidates = []
    for url in source_urls:
        key = f"{query_name}_{hashlib.sha1(url.encode('utf-8')).hexdigest()[:12]}"
        candidates.append((url, key))
    direct = direct_baidu_url(query_name)
    if direct not in source_urls:
        candidates.append((direct, query_name))
    return candidates


def clean_text(value: str) -> str:
    value = html.unescape(value or "")
    value = re.sub(r"\[[\d,\-\s]+\]", "", value)
    value = re.sub(r"\s+", " ", value)
    value = value.replace("\xa0", " ").strip()
    return value


def extract_meta(soup: BeautifulSoup, name: str) -> str:
    tag = soup.find("meta", attrs={"name": name})
    return clean_text(tag.get("content", "")) if tag else ""


def extract_canonical(soup: BeautifulSoup) -> str:
    tag = soup.find("link", attrs={"rel": "canonical"})
    return tag.get("href", "") if tag else ""


def find_section_blocks(soup: BeautifulSoup, section_name: str) -> list[str]:
    header = soup.find(attrs={"data-name": section_name})
    if not header:
        return []

    try:
        header_level = int(header.get("data-level") or 1)
    except ValueError:
        header_level = 1

    blocks = []
    for sibling in header.find_next_siblings():
        if sibling.get("data-tag") == "header":
            try:
                sibling_level = int(sibling.get("data-level") or header_level)
            except ValueError:
                sibling_level = header_level
            if sibling_level <= header_level:
                break
            continue
        if sibling.get("data-tag") != "paragraph":
            continue
        text = clean_text(sibling.get_text(" ", strip=True))
        if text:
            blocks.append(text)
    return blocks


def parse_baidu_html(chinese_name: str, html_text: str) -> ParsedPage:
    soup = BeautifulSoup(html_text, "html.parser")
    page_title = clean_text(soup.title.get_text(" ", strip=True) if soup.title else "")
    canonical = extract_canonical(soup)
    date_updated = extract_meta(soup, "dateUpdate")
    summary = extract_meta(soup, "description")
    timeline = []
    for section_name in ["人物履历", "人物经历", "个人履历", "工作履历", "生平", "工作经历", "职业经历", "学术背景与工作经历"]:
        timeline = find_section_blocks(soup, section_name)
        if timeline:
            break

    notes = []
    if "百度安全验证" in page_title or "百度安全验证" in html_text[:3000]:
        notes.append("Baidu returned security verification page")
    if not timeline:
        notes.append("No 人物履历 section extracted")

    generic_summary = summary.startswith("百度百科是一部内容开放")
    if generic_summary and not timeline:
        status = "matched_empty_baidu_page"
        notes.append("Baidu page appears to be an empty duplicate/placeholder")
    elif chinese_name and chinese_name in page_title:
        status = "matched_title"
    elif chinese_name and chinese_name in summary:
        status = "matched_summary"
    elif timeline:
        status = "timeline_found_name_uncertain"
    else:
        status = "not_matched"

    url = canonical or ("https://baike.baidu.com/item/" + urllib.parse.quote(chinese_name))
    return ParsedPage(
        url=url,
        canonical_url=canonical,
        title=page_title,
        date_updated=date_updated,
        summary=summary,
        timeline=timeline,
        match_status=status,
        notes="; ".join(notes),
    )


def protect_known_ambiguous_match(raw_name: str, parsed: ParsedPage) -> ParsedPage:
    military_markers = ["解放军", "军区", "战区", "集团军", "军委", "空军", "陆军", "海军", "火箭军", "武警"]
    if "解放军" in raw_name and not any(marker in parsed.summary for marker in military_markers):
        parsed.timeline = []
        parsed.summary = ""
        parsed.match_status = "ambiguous_name_no_baidu_match"
        parsed.notes = "Base-name Baidu page is not the PLA Wang Kai; Baidu Baike match not used"
    return parsed


def is_usable(parsed: ParsedPage) -> bool:
    if parsed.match_status in {"fetch_error", "not_matched", "matched_empty_baidu_page", "ambiguous_name_no_baidu_match"}:
        return False
    return bool(parsed.timeline or parsed.summary)


def parse_period(text: str) -> tuple[str, str, str, str]:
    compact = text.strip()
    patterns = [
        r"(?P<start>\d{4})(?:[.年](?P<start_month>\d{1,2}))?\s*年?\s*[-—－–~至]+\s*(?P<end>\d{4}|现在|今|至今)?(?:[.年](?P<end_month>\d{1,2}))?\s*年?\s*(?P<role>.*)$",
        r"(?P<start>\d{4})年\s*(?P<role>.*)$",
    ]
    for pattern in patterns:
        match = re.search(pattern, compact)
        if match:
            start = match.groupdict().get("start") or ""
            end = match.groupdict().get("end") or start
            if end in {"现在", "今", "至今"}:
                end = ""
            role = match.groupdict().get("role", "").strip()
            role = re.sub(r"^[，,、：:\s]+", "", role)
            return start, end, role, compact
    return "", "", compact, compact


def detect_systems(role_text: str) -> str:
    hits = []
    for tag, keywords in SYSTEM_KEYWORDS.items():
        if any(keyword in role_text for keyword in keywords):
            hits.append(tag)
    return "; ".join(hits)


def is_place_name_inside_institution(text: str, start: int, end: int) -> bool:
    following = text[end : end + 4]
    return any(following.startswith(suffix) for suffix in PLACE_NAME_INSTITUTION_SUFFIXES)


def detect_places(role_text: str) -> str:
    anchors = []
    for term in INSTITUTION_TERMS:
        if term in role_text and term not in anchors:
            anchors.append(term)

    for place in PROVINCE_TERMS:
        for match in re.finditer(re.escape(place), role_text):
            if is_place_name_inside_institution(role_text, match.start(), match.end()):
                continue
            if place not in anchors:
                anchors.append(place)
            break
    return "; ".join(anchors)


def existing_headers(ws) -> dict[str, int]:
    return {str(ws.cell(1, col).value).strip(): col for col in range(1, ws.max_column + 1) if ws.cell(1, col).value}


def ensure_columns(ws, headers: list[str]) -> dict[str, int]:
    header_map = existing_headers(ws)
    col = ws.max_column
    for header in headers:
        if header not in header_map:
            col += 1
            ws.cell(1, col).value = header
            header_map[header] = col
    return header_map


def reset_sheet(wb, name: str):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def autosize(ws, max_width: int = 80):
    for column_cells in ws.columns:
        letter = column_cells[0].column_letter
        width = 10
        for cell in column_cells[:200]:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(max_width, len(value) + 2))
        ws.column_dimensions[letter].width = width


def split_tags(value: str) -> set[str]:
    return {item.strip() for item in str(value or "").split(";") if item.strip()}


def parse_pbsc_faction_workbook() -> dict[str, list[dict]]:
    if not FACTION_WORKBOOK.exists():
        return {}

    faction_wb = openpyxl.load_workbook(FACTION_WORKBOOK, read_only=True, data_only=True)
    assignments: dict[str, list[dict]] = {}
    for sheet in faction_wb.worksheets:
        pbsc_pinyin = sheet.title.strip()
        pbsc_chinese = PBSC_PINYIN_TO_CN.get(pbsc_pinyin, pbsc_pinyin)
        headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        header_map = {header: idx for idx, header in enumerate(headers) if header}
        note_idx = header_map.get("Faction", header_map.get("Overlap"))
        title_idx = header_map.get("Title")
        branch_idx = header_map.get("Branch of government")

        for row in sheet.iter_rows(min_row=2, values_only=True):
            values = ["" if value is None else str(value).strip() for value in row]
            if not any(values):
                continue
            first_two = values[:2]
            cjk_values = [value for value in first_two if contains_cjk(value)]
            latin_values = [value for value in first_two if value and not contains_cjk(value)]
            if not cjk_values:
                continue
            chinese_name = normalize_name(cjk_values[0])
            pinyin_name = latin_values[0] if latin_values else ""
            note = values[note_idx] if note_idx is not None and note_idx < len(values) else ""
            title = values[title_idx] if title_idx is not None and title_idx < len(values) else ""
            branch = values[branch_idx] if branch_idx is not None and branch_idx < len(values) else ""
            named_idxs = {idx for idx in [note_idx, title_idx, branch_idx, 0, 1] if idx is not None}
            extras = [value for idx, value in enumerate(values) if idx not in named_idxs and value]
            assignment = {
                "person_chinese": chinese_name,
                "person_pinyin": pinyin_name,
                "pbsc_pinyin": pbsc_pinyin,
                "pbsc_chinese": pbsc_chinese,
                "faction_note": note,
                "title": title,
                "branch": branch,
                "extra_notes": "; ".join(extras),
                "source_sheet": sheet.title,
            }
            assignments.setdefault(chinese_name, []).append(assignment)
    return assignments


def year_interval(start, end):
    try:
        start_year = int(start)
    except (TypeError, ValueError):
        return None
    try:
        end_year = int(end) if end else CURRENT_YEAR
    except (TypeError, ValueError):
        end_year = CURRENT_YEAR
    if end_year < start_year:
        end_year = start_year
    return start_year, end_year


def intervals_overlap(left, right) -> bool:
    if not left or not right:
        return False
    return left[0] <= right[1] and right[0] <= left[1]


def overlap_window(left, right):
    return max(left[0], right[0]), min(left[1], right[1])


def anchor_segments(text: str, anchor: str) -> list[str]:
    parts = re.split(r"[；;。]|[（）()]", text or "")
    return [part.strip() for part in parts if anchor in part]


def anchor_interval(anchor: str, raw_text: str, default_interval):
    for segment in anchor_segments(raw_text, anchor):
        start, end, _role, _raw = parse_period(segment)
        interval = year_interval(start, end)
        if interval:
            return interval
    return default_interval


def episode_anchor_records(ep: dict) -> list[dict]:
    records = []
    seen = set()
    for anchor in sorted(ep["places"], key=len, reverse=True):
        if anchor in EXCLUDED_OVERLAP_ANCHORS:
            continue
        interval = anchor_interval(anchor, ep["raw"], ep["interval"])
        key = (anchor, interval)
        if key in seen:
            continue
        seen.add(key)
        records.append({"anchor": anchor, "interval": interval})
    return records


def build_pbsc_pair_sheet(wb, episode_rows: list[list], assignments_by_name: dict[str, list[dict]], title_by_name: dict[str, str]):
    ws = reset_sheet(wb, PBSC_PAIR_SHEET)
    ws.append([
        "Person Chinese Name",
        "Person Pinyin Name",
        "Person Title",
        "PBSC Chinese Name",
        "PBSC Pinyin Name",
        "Explicit Faction Workbook Match",
        "Faction Workbook Note",
        "Strict Overlap Score",
        "Physical Overlap Pairs",
        "Excluded Broad-Tag Pairs",
        "Shared Physical Anchors",
        "Context System Tags",
        "Earliest Overlap Start",
        "Latest Overlap End",
        "Sample Overlap Evidence (CN)",
    ])

    episodes = []
    for row in episode_rows:
        interval = year_interval(row[4], row[5])
        episodes.append({
            "pinyin": row[0],
            "name": row[1],
            "is_pbsc": row[2] == "Y",
            "episode_no": row[3],
            "interval": interval,
            "start": row[4],
            "end": row[5],
            "role": row[6],
            "raw": row[7],
            "places": split_tags(row[8]),
            "systems": split_tags(row[9]),
            "anchor_records": [],
        })
    for ep in episodes:
        ep["anchor_records"] = episode_anchor_records(ep)

    pbsc_episodes = [ep for ep in episodes if ep["is_pbsc"] and ep["interval"]]
    people = sorted({(ep["name"], ep["pinyin"]) for ep in episodes if not ep["is_pbsc"]})

    for person_name, person_pinyin in people:
        person_eps = [ep for ep in episodes if ep["name"] == person_name and ep["interval"]]
        for pbsc_name in sorted(KNOWN_PBSC_CN):
            faction_hits = [
                item for item in assignments_by_name.get(normalize_name(person_name), [])
                if item["pbsc_chinese"] == pbsc_name
            ]
            faction_hit_note = format_faction_matches(faction_hits) if faction_hits else ""
            pbsc_eps = [ep for ep in pbsc_episodes if ep["name"] == pbsc_name]
            if not pbsc_eps:
                continue
            pbsc_pinyin = pbsc_eps[0]["pinyin"]
            shared_anchors = set()
            shared_systems = set()
            evidence = []
            overlap_windows = []
            pair_count = 0
            weak_pair_count = 0
            for left in person_eps:
                for right in pbsc_eps:
                    if not intervals_overlap(left["interval"], right["interval"]):
                        continue
                    systems = left["systems"] & right["systems"]
                    matching_anchor_windows = []
                    for left_anchor in left["anchor_records"]:
                        for right_anchor in right["anchor_records"]:
                            if left_anchor["anchor"] != right_anchor["anchor"]:
                                continue
                            if not intervals_overlap(left_anchor["interval"], right_anchor["interval"]):
                                continue
                            matching_anchor_windows.append((left_anchor["anchor"], overlap_window(left_anchor["interval"], right_anchor["interval"])))
                    if not matching_anchor_windows:
                        if systems:
                            weak_pair_count += 1
                        continue
                    pair_count += 1
                    shared_systems.update(systems)
                    for anchor, window in matching_anchor_windows:
                        shared_anchors.add(anchor)
                        overlap_windows.append(window)
                    if len(evidence) < 5:
                        left_years = f"{left['start']}-{left['end'] or CURRENT_YEAR}"
                        right_years = f"{right['start']}-{right['end'] or CURRENT_YEAR}"
                        anchor_text = ", ".join(
                            sorted(
                                {
                                    f"{anchor} {window[0]}-{window[1]}"
                                    for anchor, window in matching_anchor_windows
                                }
                            )
                        )
                        evidence.append(
                            f"{anchor_text}: {left_years} {left['role']} ↔ {pbsc_name} {right_years} {right['role']}"
                        )
            if pair_count or faction_hits:
                score = (5 * pair_count) + (10 * len(shared_anchors)) + (3 * len(shared_systems)) + (50 if faction_hits else 0)
                earliest = min((start for start, _ in overlap_windows), default="")
                latest = max((end for _, end in overlap_windows), default="")
                ws.append([
                    person_name,
                    person_pinyin,
                    title_by_name.get(person_name, ""),
                    pbsc_name,
                    pbsc_pinyin,
                    "Y" if faction_hits else "",
                    faction_hit_note,
                    score,
                    pair_count,
                    weak_pair_count,
                    "; ".join(sorted(shared_anchors)),
                    "; ".join(sorted(shared_systems)),
                    earliest,
                    latest,
                    "\n".join(evidence),
                ])
    autosize(ws, max_width=100)
    ws.freeze_panes = "A2"
    return ws


def build_gap_sheet(wb, report: list[dict]):
    ws = reset_sheet(wb, GAP_SHEET)
    ws.append([
        "Workbook Row",
        "Pinyin Name",
        "Chinese Name",
        "Baidu Query Name",
        "Baidu Match Status",
        "Episode Count",
        "Baidu URL",
        "Notes",
    ])
    for item in report:
        if item["status"] != "matched_title" or item["episode_count"] == 0:
            ws.append([
                item["row"],
                item["pinyin_name"],
                item["chinese_name"],
                item.get("query_name", ""),
                item["status"],
                item["episode_count"],
                item["url"],
                item["notes"],
            ])
    autosize(ws, max_width=100)
    ws.freeze_panes = "A2"
    return ws


def format_faction_matches(assignments: list[dict]) -> str:
    parts = []
    for item in assignments:
        note = item["faction_note"] or item["extra_notes"] or "listed"
        parts.append(f"{item['pbsc_pinyin']}: {note}")
    return "; ".join(parts)


def format_faction_notes(assignments: list[dict]) -> str:
    rows = []
    for item in assignments:
        details = [item["faction_note"], item["extra_notes"], item["branch"], item["title"]]
        details = [detail for detail in details if detail]
        rows.append(f"{item['pbsc_pinyin']} ({item['pbsc_chinese']}): " + " | ".join(details))
    return "\n".join(rows)


def build_faction_sheet(wb, assignments_by_name: dict[str, list[dict]]):
    ws = reset_sheet(wb, FACTION_SHEET)
    ws.append([
        "Person Chinese Name",
        "Person Pinyin Name",
        "PBSC Chinese Name",
        "PBSC Pinyin Name",
        "Faction/Overlap Note",
        "Extra Notes",
        "Title From Faction Workbook",
        "Branch From Faction Workbook",
        "Source Sheet",
    ])
    for person_name in sorted(assignments_by_name):
        for item in assignments_by_name[person_name]:
            ws.append([
                item["person_chinese"],
                item["person_pinyin"],
                item["pbsc_chinese"],
                item["pbsc_pinyin"],
                item["faction_note"],
                item["extra_notes"],
                item["title"],
                item["branch"],
                item["source_sheet"],
            ])
    autosize(ws, max_width=100)
    ws.freeze_panes = "A2"
    return ws


def main():
    wb = openpyxl.load_workbook(WORKBOOK)
    ws = wb[MAIN_SHEET]
    headers = ensure_columns(ws, ADDED_COLUMNS + FACTION_COLUMNS)
    original_headers = existing_headers(ws)
    generated_cols = {headers[column] for column in ADDED_COLUMNS + FACTION_COLUMNS if column in headers}
    chinese_col = original_headers["Chinese Name"]
    pinyin_col = original_headers["Pinyin Name"]
    title_col = original_headers.get("Title")
    title_by_name = {
        str(ws.cell(row, chinese_col).value or "").strip(): str(ws.cell(row, title_col).value or "").strip()
        for row in range(2, ws.max_row + 1)
        if title_col and str(ws.cell(row, chinese_col).value or "").strip()
    }
    faction_assignments = parse_pbsc_faction_workbook()

    episode_ws = reset_sheet(wb, EPISODE_SHEET)
    episode_headers = [
        "Pinyin Name",
        "Chinese Name",
        "Is 20th PBSC",
        "Episode No.",
        "Start Year",
        "End Year",
        "Role Text (CN)",
        "Raw Timeline Entry (CN)",
        "Detected Places",
        "Detected System Tags",
        "Baidu URL",
        "Baidu Last Updated",
        "Baidu Match Status",
    ]
    episode_ws.append(episode_headers)

    overlap_ws = reset_sheet(wb, PBSC_OVERLAP_SHEET)
    overlap_ws.append([
        "Chinese Name",
        "Pinyin Name",
        "Is 20th PBSC",
        "Baidu Episode Count",
        "Detected Places Across Career",
        "Detected System Tags Across Career",
        "Baidu URL",
        "Baidu Match Status",
        "Current Title",
    ])

    report = []
    episode_rows = []
    overlap_rows = []

    for row in range(2, ws.max_row + 1):
        chinese_name = str(ws.cell(row, chinese_col).value or "").strip()
        pinyin_name = str(ws.cell(row, pinyin_col).value or "").strip()
        if not chinese_name:
            continue
        query_name = baidu_query_name(chinese_name)
        source_baidu_urls = baidu_urls_from_row(ws, row, generated_cols)

        fetch_attempts = []
        try:
            parsed = None
            source = ""
            last_exc = None
            for candidate_url, cache_key in candidate_baidu_urls(query_name, source_baidu_urls):
                try:
                    html_text, source = fetch_baidu_url(candidate_url, cache_key)
                except Exception as exc:
                    last_exc = exc
                    fetch_attempts.append({
                        "url": candidate_url,
                        "source": "error",
                        "status": "fetch_error",
                        "episode_count": 0,
                        "error": str(exc),
                    })
                    continue
                candidate = parse_baidu_html(query_name, html_text)
                candidate = protect_known_ambiguous_match(chinese_name, candidate)
                fetch_attempts.append({
                    "url": candidate_url,
                    "source": source,
                    "status": candidate.match_status,
                    "episode_count": len(candidate.timeline),
                })
                parsed = candidate
                if is_usable(candidate):
                    break
            if parsed is None:
                raise last_exc or RuntimeError("No Baidu candidates available")
        except Exception as exc:
            parsed = ParsedPage(
                url=direct_baidu_url(query_name),
                canonical_url="",
                title="",
                date_updated="",
                summary="",
                timeline=[],
                match_status="fetch_error",
                notes=str(exc),
            )
            source = "error"

        timeline_text = "\n".join(parsed.timeline)
        ws.cell(row, headers["Baidu Query Name"]).value = query_name
        ws.cell(row, headers["Baidu URL"]).value = parsed.url
        ws.cell(row, headers["Baidu Last Updated"]).value = parsed.date_updated
        ws.cell(row, headers["Baidu Profile Summary (CN)"]).value = parsed.summary
        ws.cell(row, headers["Baidu Career Timeline (CN)"]).value = timeline_text
        ws.cell(row, headers["Baidu Timeline Episode Count"]).value = len(parsed.timeline)
        ws.cell(row, headers["Baidu Match Status"]).value = parsed.match_status
        ws.cell(row, headers["Baidu Notes"]).value = parsed.notes

        normalized_person_name = normalize_name(chinese_name)
        person_factions = faction_assignments.get(normalized_person_name, [])
        ws.cell(row, headers["PBSC Faction Workbook Count"]).value = len(person_factions)
        ws.cell(row, headers["PBSC Faction Workbook Matches"]).value = format_faction_matches(person_factions)
        ws.cell(row, headers["PBSC Faction Workbook Notes"]).value = format_faction_notes(person_factions)

        all_places = set()
        all_systems = set()
        is_pbsc = "Y" if chinese_name in KNOWN_PBSC_CN else ""

        for idx, entry in enumerate(parsed.timeline, start=1):
            start_year, end_year, role_text, raw = parse_period(entry)
            places = detect_places(role_text)
            systems = detect_systems(role_text)
            all_places.update(filter(None, [p.strip() for p in places.split(";")]))
            all_systems.update(filter(None, [s.strip() for s in systems.split(";")]))
            episode_rows.append([
                pinyin_name,
                chinese_name,
                is_pbsc,
                idx,
                start_year,
                end_year,
                role_text,
                raw,
                places,
                systems,
                parsed.url,
                parsed.date_updated,
                parsed.match_status,
            ])

        overlap_rows.append([
            chinese_name,
            pinyin_name,
            is_pbsc,
            len(parsed.timeline),
            "; ".join(sorted(all_places)),
            "; ".join(sorted(all_systems)),
            parsed.url,
            parsed.match_status,
            str(ws.cell(row, title_col).value or "") if title_col else "",
        ])

        report.append({
            "row": row,
            "chinese_name": chinese_name,
            "query_name": query_name,
            "pinyin_name": pinyin_name,
            "source": source,
            "attempts": fetch_attempts,
            "status": parsed.match_status,
            "episode_count": len(parsed.timeline),
            "url": parsed.url,
            "notes": parsed.notes,
        })

        if (row - 1) % 25 == 0:
            print(f"Processed {row - 1} people")

    for values in episode_rows:
        episode_ws.append(values)
    for values in overlap_rows:
        overlap_ws.append(values)

    for target in [ws, episode_ws, overlap_ws]:
        autosize(target)
    pair_ws = build_pbsc_pair_sheet(wb, episode_rows, faction_assignments, title_by_name)
    gap_ws = build_gap_sheet(wb, report)
    faction_ws = build_faction_sheet(wb, faction_assignments)
    ws.freeze_panes = "A2"
    episode_ws.freeze_panes = "A2"
    overlap_ws.freeze_panes = "A2"

    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    wb.save(WORKBOOK)

    statuses = {}
    for item in report:
        statuses[item["status"]] = statuses.get(item["status"], 0) + 1
    print(json.dumps({
        "workbook": str(WORKBOOK),
        "people": len(report),
        "statuses": statuses,
        "episodes": len(episode_rows),
        "pbsc_overlap_candidate_rows": pair_ws.max_row - 1,
        "gap_review_rows": gap_ws.max_row - 1,
        "pbsc_faction_assignment_rows": faction_ws.max_row - 1,
        "report": str(REPORT),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
