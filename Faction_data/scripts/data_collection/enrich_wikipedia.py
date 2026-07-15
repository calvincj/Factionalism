"""
Wikipedia cross-check for the Baidu-derived career data.

For every person in the "20th Central Committee" sheet, fetches the Chinese
Wikipedia and English Wikipedia articles (if any), extracts structured
office/term_start/term_end entries from the infobox, and writes a workbook
that lets a human line up Wikipedia's version of each person's career against
what enrich_baidu.py pulled from Baidu Baike.

This does NOT auto-resolve conflicts between the two sources -- it surfaces
them side by side (Wikipedia coverage where Baidu has none, and Wikipedia's
office/date list next to Baidu's episode list) so a human can judge whether
they agree. Treat "Wikipedia has an infobox" as a second opinion, not ground
truth: Wikipedia editors also make mistakes, and politically sensitive
tenures are exactly where both sources can be wrong in different ways.

Usage:
    python3 enrich_wikipedia.py                # run on everyone
    python3 enrich_wikipedia.py 何立峰 李桥铭    # run on just these people (testing)
"""

import json
import re
import sys
import time
from pathlib import Path

import mwparserfromhell
import openpyxl
import requests

sys.path.insert(0, str(Path(__file__).resolve().parent))
from enrich_baidu import autosize, contains_cjk, existing_headers  # noqa: E402

SCRIPT_DIR = Path(__file__).resolve().parent
FACTION_DATA_DIR = SCRIPT_DIR.parent.parent  # scripts/data_collection/ -> scripts/ -> Faction_data/
SOURCE_WORKBOOK = FACTION_DATA_DIR / "source" / "Chinese Leadership Database - Baidu enriched working copy.xlsx"
OUT_XLSX = FACTION_DATA_DIR / "outputs" / "PBSC Wikipedia cross-check.xlsx"
CACHE_DIR = FACTION_DATA_DIR / "scrape_cache" / "wikipedia_cache"

MAIN_SHEET = "20th Central Committee"
EPISODE_SHEET = "Baidu Career Episodes"

API_URL = {
    "zh": "https://zh.wikipedia.org/w/api.php",
    "en": "https://en.wikipedia.org/w/api.php",
}

HEADERS = {
    "User-Agent": (
        "BrookingsFactionalismResearch/0.1 "
        "(academic research project on CCP elite career-overlap data validation; "
        "contact calvinc7028@gmail.com)"
    )
}

POLITICAL_SIGNAL_TERMS = [
    "政治人物", "中国共产党", "中华人民共和国", "中央委员", "省委书记", "省长", "部长",
    "Chinese politician", "Communist Party", "Politburo", "State Council", "provincial",
    "PLA", "People's Liberation Army", "解放军",
]

REQUEST_DELAY_SECONDS = 0.25
SESSION = requests.Session()


def cache_path(lang, title):
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    slug = re.sub(r"[^\w一-鿿-]+", "_", title).strip("_")
    if len(slug) > 80:
        slug = slug[:80]
    return CACHE_DIR / f"{lang}_{slug}.json"


def api_get(lang, title):
    """Fetch (and cache) raw MediaWiki API JSON for one title. Returns parsed JSON or None."""
    path = cache_path(lang, title)
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))

    params = {
        "action": "query",
        "format": "json",
        "titles": title,
        "redirects": 1,
        "prop": "revisions|extracts|pageprops",
        "rvprop": "content",
        "rvslots": "main",
        "exintro": 1,
        "explaintext": 1,
    }
    try:
        resp = SESSION.get(API_URL[lang], params=params, headers=HEADERS, timeout=20)
        resp.raise_for_status()
        data = resp.json()
    except (requests.RequestException, ValueError) as exc:
        data = {"error": str(exc)}
    time.sleep(REQUEST_DELAY_SECONDS)
    path.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
    return data


def extract_page(data):
    """Pull the single page dict out of a query response. None if missing/error."""
    if not data or "error" in data:
        return None
    pages = data.get("query", {}).get("pages", {})
    for pid, page in pages.items():
        if pid == "-1" or "missing" in page:
            return None
        wikitext = page.get("revisions", [{}])[0].get("slots", {}).get("main", {}).get("*", "")
        return {
            "title": page.get("title", ""),
            "wikitext": wikitext,
            "extract": page.get("extract", "") or "",
            "pageprops": page.get("pageprops", {}) or {},
        }
    return None


def is_disambiguation(page):
    return "disambiguation" in page.get("pageprops", {})


def political_signal_score(text):
    return sum(1 for term in POLITICAL_SIGNAL_TERMS if term in text)


def birth_year_from_extract(extract):
    # Chinese Wikipedia bios almost always open "姓名（1953年6月15日—...）" right after the
    # name -- check only the opening of the text so we don't pick up unrelated years
    # (career dates, other people's birthdates) mentioned later in a long extract.
    head = extract[:120]
    match = re.search(r"[（(](\d{4})年", head)
    if match:
        return int(match.group(1))
    match = re.search(r"\bborn[^\d]{0,20}(\d{4})", head, re.IGNORECASE)
    if match:
        return int(match.group(1))
    match = re.search(r"(\d{4})年\d{1,2}月\d{1,2}日.{0,6}出生", head)
    if match:
        return int(match.group(1))
    return None


def disambiguation_candidates(lang, page):
    parsed = mwparserfromhell.parse(page["wikitext"])
    titles = []
    for link in parsed.filter_wikilinks():
        target = str(link.title).strip()
        if not target or target.startswith(("Category:", "File:", "分类:", "文件:")):
            continue
        if target not in titles:
            titles.append(target)
    return titles[:8]


def resolve(lang, title, expected_birth_year, hint_terms):
    """Fetch a title; if it's a disambiguation page, try to pick the right candidate.

    Returns (page_dict_or_None, status, note).
    """
    data = api_get(lang, title)
    page = extract_page(data)
    if page is None:
        return None, "not_found", ""

    if not is_disambiguation(page):
        score = political_signal_score(page["extract"] + " " + " ".join(hint_terms))
        found_year = birth_year_from_extract(page["extract"])
        year_ok = expected_birth_year is None or found_year is None or found_year == expected_birth_year
        if score == 0 and not year_ok:
            return page, "matched_low_confidence", f"no political-signal terms; birth year {found_year} vs expected {expected_birth_year}"
        if not year_ok:
            return page, "matched_low_confidence", f"birth year mismatch: page says {found_year}, source says {expected_birth_year}"
        return page, "matched", ""

    # Disambiguation page: try each candidate, score, keep the best.
    best = None
    best_score = -1
    best_title = ""
    for candidate_title in disambiguation_candidates(lang, page):
        cand_data = api_get(lang, candidate_title)
        cand_page = extract_page(cand_data)
        if cand_page is None or is_disambiguation(cand_page):
            continue
        score = political_signal_score(cand_page["extract"])
        found_year = birth_year_from_extract(cand_page["extract"])
        if expected_birth_year is not None and found_year == expected_birth_year:
            score += 5
        if score > best_score:
            best_score = score
            best = cand_page
            best_title = candidate_title
    if best is not None and best_score > 0:
        return best, "matched_via_disambiguation", f"resolved to '{best_title}' (signal score {best_score})"
    return None, "disambiguation_unresolved", f"{len(disambiguation_candidates(lang, page))} candidates, none confidently matched"


def parse_infobox_offices(wikitext):
    parsed = mwparserfromhell.parse(wikitext)
    infobox = None
    for template in parsed.filter_templates():
        if "infobox" in str(template.name).strip().lower():
            infobox = template
            break
    if infobox is None:
        return []

    slots = {}
    for param in infobox.params:
        name = str(param.name).strip()
        value = param.value.strip_code().strip()
        if not value:
            continue
        for key in ("office", "title", "term_start", "term_end"):
            if name == key or (name.startswith(key) and name[len(key):].isdigit()):
                idx = name[len(key):] or "0"
                field = "office" if key == "title" else key
                slots.setdefault(idx, {})[field] = value

    offices = []
    for idx in sorted(slots, key=lambda x: int(x)):
        entry = slots[idx]
        if entry.get("office"):
            offices.append({
                "office": entry.get("office", ""),
                "term_start": entry.get("term_start", ""),
                "term_end": entry.get("term_end", ""),
            })
    return offices


def format_offices(offices):
    return "; ".join(
        f"{o['office']} ({o['term_start']}-{o['term_end'] or 'present'})" if (o["term_start"] or o["term_end"]) else o["office"]
        for o in offices
    )


def load_source_people():
    wb = openpyxl.load_workbook(SOURCE_WORKBOOK, read_only=True, data_only=True)
    ws = wb[MAIN_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(x or "").strip() for x in rows[0]]
    idx = {h: i for i, h in enumerate(headers) if h}
    people = []
    for row in rows[1:]:
        chinese_name = row[idx["Chinese Name"]]
        if not chinese_name:
            continue
        birthdate = row[idx.get("Birthdate", -1)] if "Birthdate" in idx else None
        birth_year = None
        if hasattr(birthdate, "year"):
            birth_year = birthdate.year
        elif birthdate:
            m = re.search(r"(\d{4})", str(birthdate))
            if m:
                birth_year = int(m.group(1))
        people.append({
            "chinese_name": re.sub(r"[（(].*?[）)]", "", str(chinese_name)).strip(),
            "pinyin_name": row[idx.get("Pinyin Name", -1)] if "Pinyin Name" in idx else "",
            "title": row[idx.get("Title", -1)] if "Title" in idx else "",
            "baidu_status": row[idx.get("Baidu Match Status", -1)] if "Baidu Match Status" in idx else "",
            "baidu_episode_count": row[idx.get("Baidu Timeline Episode Count", -1)] if "Baidu Timeline Episode Count" in idx else 0,
            "birth_year": birth_year,
        })
    wb.close()
    return people


def load_baidu_episode_counts_by_name():
    """Independent recount straight from Baidu Career Episodes, in case the summary column drifted."""
    wb = openpyxl.load_workbook(SOURCE_WORKBOOK, read_only=True, data_only=True)
    if EPISODE_SHEET not in wb.sheetnames:
        return {}
    ws = wb[EPISODE_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(x or "").strip() for x in rows[0]]
    name_idx = headers.index("Chinese Name") if "Chinese Name" in headers else 1
    counts = {}
    for row in rows[1:]:
        name = row[name_idx]
        if name:
            counts[name] = counts.get(name, 0) + 1
    wb.close()
    return counts


def process_person(person):
    hint_terms = [str(person.get("title") or "")]
    zh_page, zh_status, zh_note = resolve("zh", person["chinese_name"], person["birth_year"], hint_terms)
    zh_offices = parse_infobox_offices(zh_page["wikitext"]) if zh_page else []

    en_name = str(person.get("pinyin_name") or "").strip()
    en_page, en_status, en_note = (None, "no_pinyin_name", "") if not en_name else resolve("en", en_name, person["birth_year"], hint_terms)
    en_offices = parse_infobox_offices(en_page["wikitext"]) if en_page else []

    return {
        "Chinese Name": person["chinese_name"],
        "Pinyin Name": person["pinyin_name"],
        "Title": person["title"],
        "Baidu Match Status": person["baidu_status"],
        "Baidu Episode Count": person["baidu_episode_count"],
        "Wikipedia ZH Status": zh_status,
        "Wikipedia ZH Title": zh_page["title"] if zh_page else "",
        "Wikipedia ZH Note": zh_note,
        "Wikipedia ZH Offices": format_offices(zh_offices),
        "Wikipedia ZH Office Count": len(zh_offices),
        "Wikipedia ZH Extract": (zh_page["extract"][:400] if zh_page else ""),
        "Wikipedia EN Status": en_status,
        "Wikipedia EN Title": en_page["title"] if en_page else "",
        "Wikipedia EN Note": en_note,
        "Wikipedia EN Offices": format_offices(en_offices),
        "Wikipedia EN Office Count": len(en_offices),
        "Wikipedia EN Extract": (en_page["extract"][:400] if en_page else ""),
    }


def classify_row(row):
    baidu_has_data = int(row["Baidu Episode Count"] or 0) > 0
    wiki_has_data = row["Wikipedia ZH Office Count"] > 0 or row["Wikipedia EN Office Count"] > 0
    if not baidu_has_data and wiki_has_data:
        return "gap_filled_by_wikipedia"
    if baidu_has_data and wiki_has_data:
        return "both_sources_have_data_review_for_consistency"
    if baidu_has_data and not wiki_has_data:
        return "baidu_only"
    return "neither_source_has_structured_data"


def main():
    filter_names = set(sys.argv[1:]) or None

    people = load_source_people()
    baidu_counts = load_baidu_episode_counts_by_name()
    if filter_names:
        people = [p for p in people if p["chinese_name"] in filter_names]

    print(f"Processing {len(people)} people...")
    rows = []
    for i, person in enumerate(people, 1):
        person["baidu_episode_count"] = baidu_counts.get(person["chinese_name"], person["baidu_episode_count"] or 0)
        row = process_person(person)
        row["Cross-Check Flag"] = classify_row(row)
        rows.append(row)
        if i % 25 == 0 or i == len(people):
            print(f"  {i}/{len(people)}")

    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    ws = out_wb.create_sheet("Wikipedia Cross-Check")
    headers = list(rows[0].keys()) if rows else []
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    ws.freeze_panes = "A2"
    autosize(ws, max_width=90)

    from collections import Counter
    flag_counts = Counter(r["Cross-Check Flag"] for r in rows)
    zh_status_counts = Counter(r["Wikipedia ZH Status"] for r in rows)
    en_status_counts = Counter(r["Wikipedia EN Status"] for r in rows)

    summary_ws = out_wb.create_sheet("Summary")
    summary_ws.append(["Metric", "Value"])
    summary_ws.append(["People processed", len(rows)])
    for flag, count in flag_counts.most_common():
        summary_ws.append([f"cross_check_flag: {flag}", count])
    for status, count in zh_status_counts.most_common():
        summary_ws.append([f"zh_status: {status}", count])
    for status, count in en_status_counts.most_common():
        summary_ws.append([f"en_status: {status}", count])
    autosize(summary_ws)

    gap_ws = out_wb.create_sheet("Baidu Gaps Filled")
    gap_rows = [r for r in rows if r["Cross-Check Flag"] == "gap_filled_by_wikipedia"]
    if gap_rows:
        gap_ws.append(headers)
        for row in gap_rows:
            gap_ws.append([row.get(h, "") for h in headers])
        autosize(gap_ws, max_width=90)

    OUT_XLSX.parent.mkdir(exist_ok=True)
    out_wb.save(OUT_XLSX)

    print(json.dumps({
        "cross_check_flags": dict(flag_counts),
        "zh_status": dict(zh_status_counts),
        "en_status": dict(en_status_counts),
        "output": str(OUT_XLSX),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
