import json
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl


ROOT = Path(r"C:\Users\AMatthias\Documents\Codex\2026-05-12\i-want-to-make-database-that")
WORKBOOK = ROOT / "Chinese Leadership Database - Baidu enriched working copy.xlsx"
OUT_MD = ROOT / "PBSC overlap analysis report - with titles.md"
OUT_XLSX = ROOT / "PBSC overlap analysis tables - with titles.xlsx"

PROVINCES = {
    "北京", "天津", "上海", "重庆", "河北", "山西", "辽宁", "吉林", "黑龙江", "江苏", "浙江", "安徽",
    "福建", "江西", "山东", "河南", "湖北", "湖南", "广东", "海南", "四川", "贵州", "云南", "陕西",
    "甘肃", "青海", "台湾", "内蒙古", "广西", "西藏", "宁夏", "新疆", "香港", "澳门",
}


def rows_as_dicts(ws):
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(x or "").strip() for x in rows[0]]
    out = []
    for values in rows[1:]:
        item = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        out.append(item)
    return out


def split_semicolon(value):
    return [item.strip() for item in str(value or "").split(";") if item and item.strip()]


def anchor_type(anchor):
    return "province/place" if anchor in PROVINCES else "institution/org"


def make_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    for col in ws.columns:
        width = 10
        for cell in col[:200]:
            width = max(width, min(80, len(str(cell.value or "")) + 2))
        ws.column_dimensions[col[0].column_letter].width = width


def main():
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    main_rows = rows_as_dicts(wb["20th Central Committee"])
    title_by_name = {r.get("Chinese Name"): r.get("Title") for r in main_rows if r.get("Chinese Name")}
    candidate_rows = rows_as_dicts(wb["PBSC Overlap Candidates"])
    faction_rows = rows_as_dicts(wb["PBSC Faction Assignments"])
    gap_rows = rows_as_dicts(wb["Baidu Gap Review"])
    episode_rows = rows_as_dicts(wb["Baidu Career Episodes"])

    physical_rows = [r for r in candidate_rows if r.get("Shared Physical Anchors")]
    manual_rows = [r for r in candidate_rows if r.get("Explicit Faction Workbook Match") == "Y"]
    manual_no_physical = [r for r in manual_rows if not r.get("Shared Physical Anchors")]

    by_pbsc = {}
    for pbsc in sorted({r["PBSC Pinyin Name"] for r in candidate_rows if r.get("PBSC Pinyin Name")}):
        rows = [r for r in candidate_rows if r.get("PBSC Pinyin Name") == pbsc]
        phys = [r for r in rows if r.get("Shared Physical Anchors")]
        manual = [r for r in rows if r.get("Explicit Faction Workbook Match") == "Y"]
        anchors = Counter()
        place_anchors = Counter()
        institution_anchors = Counter()
        for r in phys:
            for anchor in split_semicolon(r.get("Shared Physical Anchors")):
                anchors[anchor] += 1
                if anchor_type(anchor) == "province/place":
                    place_anchors[anchor] += 1
                else:
                    institution_anchors[anchor] += 1
        by_pbsc[pbsc] = {
            "pbsc_chinese": next((r["PBSC Chinese Name"] for r in rows if r.get("PBSC Chinese Name")), ""),
            "candidate_people": len({r["Person Chinese Name"] for r in rows}),
            "physical_people": len({r["Person Chinese Name"] for r in phys}),
            "manual_people": len({r["Person Chinese Name"] for r in manual}),
            "physical_pairs": sum(int(r.get("Physical Overlap Pairs") or 0) for r in phys),
            "top_anchors": anchors.most_common(8),
            "top_places": place_anchors.most_common(5),
            "top_institutions": institution_anchors.most_common(5),
        }

    by_person = defaultdict(lambda: {
        "pinyin": "",
        "pbscs": set(),
        "manual_pbscs": set(),
        "score": 0,
        "pairs": 0,
        "anchors": Counter(),
    })
    for r in candidate_rows:
        person = r.get("Person Chinese Name")
        if not person:
            continue
        rec = by_person[person]
        rec["pinyin"] = r.get("Person Pinyin Name") or ""
        if r.get("Shared Physical Anchors"):
            rec["pbscs"].add(r.get("PBSC Pinyin Name"))
            rec["score"] += int(r.get("Strict Overlap Score") or 0)
            rec["pairs"] += int(r.get("Physical Overlap Pairs") or 0)
            for anchor in split_semicolon(r.get("Shared Physical Anchors")):
                rec["anchors"][anchor] += 1
        if r.get("Explicit Faction Workbook Match") == "Y":
            rec["manual_pbscs"].add(r.get("PBSC Pinyin Name"))

    top_connectors = sorted(
        [
            {
                "person": person,
                "pinyin": rec["pinyin"],
                "title": title_by_name.get(person, ""),
                "pbsc_count": len(rec["pbscs"]),
                "manual_count": len(rec["manual_pbscs"]),
                "score": rec["score"],
                "pairs": rec["pairs"],
                "anchors": "; ".join([a for a, _ in rec["anchors"].most_common(8)]),
                "pbscs": "; ".join(sorted(rec["pbscs"])),
                "manual_pbscs": "; ".join(sorted(rec["manual_pbscs"])),
            }
            for person, rec in by_person.items()
            if rec["pbscs"] or rec["manual_pbscs"]
        ],
        key=lambda x: (x["pbsc_count"], x["score"], x["pairs"]),
        reverse=True,
    )

    top_pair_rows = sorted(
        physical_rows,
        key=lambda r: (
            int(r.get("Strict Overlap Score") or 0),
            int(r.get("Physical Overlap Pairs") or 0),
        ),
        reverse=True,
    )

    anchor_counts = Counter()
    anchor_pair_counts = Counter()
    for r in physical_rows:
        for anchor in split_semicolon(r.get("Shared Physical Anchors")):
            anchor_counts[anchor] += 1
            anchor_pair_counts[anchor] += int(r.get("Physical Overlap Pairs") or 0)

    manual_keys = {(r["Person Chinese Name"], r["PBSC Pinyin Name"]) for r in manual_rows}
    physical_keys = {(r["Person Chinese Name"], r["PBSC Pinyin Name"]) for r in physical_rows}
    manual_with_physical = sorted(manual_keys & physical_keys)
    manual_without_physical = sorted(manual_keys - physical_keys)
    physical_without_manual = sorted(physical_keys - manual_keys)

    summary = {
        "central_committee_rows": len([r for r in main_rows if r.get("Chinese Name")]),
        "baidu_episode_rows": len(episode_rows),
        "baidu_gap_rows": len(gap_rows),
        "candidate_rows_total": len(candidate_rows),
        "physical_candidate_rows": len(physical_rows),
        "manual_candidate_rows": len(manual_rows),
        "manual_without_physical_rows": len(manual_no_physical),
        "unique_people_with_physical_overlap": len({r["Person Chinese Name"] for r in physical_rows}),
        "manual_assignments": len(faction_rows),
    }

    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)
    make_sheet(
        out_wb,
        "Summary",
        ["Metric", "Value"],
        [[k, v] for k, v in summary.items()],
    )
    make_sheet(
        out_wb,
        "By PBSC",
        ["PBSC", "Chinese", "Physical People", "Manual People", "Physical Pairs", "Top Places", "Top Institutions"],
        [
            [
                pbsc,
                info["pbsc_chinese"],
                info["physical_people"],
                info["manual_people"],
                info["physical_pairs"],
                "; ".join(f"{a} ({c})" for a, c in info["top_places"]),
                "; ".join(f"{a} ({c})" for a, c in info["top_institutions"]),
            ]
            for pbsc, info in by_pbsc.items()
        ],
    )
    make_sheet(
        out_wb,
        "Top Connectors",
        ["Person", "Pinyin", "Title", "PBSC Count", "Manual Count", "Score", "Physical Pairs", "PBSCs", "Manual PBSCs", "Top Anchors"],
        [
            [r["person"], r["pinyin"], r["title"], r["pbsc_count"], r["manual_count"], r["score"], r["pairs"], r["pbscs"], r["manual_pbscs"], r["anchors"]]
            for r in top_connectors
        ],
    )
    make_sheet(
        out_wb,
        "Top Pair Evidence",
        ["Person", "Pinyin", "Title", "PBSC", "PBSC Chinese", "Manual", "Score", "Pairs", "Anchors", "Start", "End", "Evidence"],
        [
            [
                r.get("Person Chinese Name"),
                r.get("Person Pinyin Name"),
                title_by_name.get(r.get("Person Chinese Name"), ""),
                r.get("PBSC Pinyin Name"),
                r.get("PBSC Chinese Name"),
                r.get("Explicit Faction Workbook Match"),
                r.get("Strict Overlap Score"),
                r.get("Physical Overlap Pairs"),
                r.get("Shared Physical Anchors"),
                r.get("Earliest Overlap Start"),
                r.get("Latest Overlap End"),
                r.get("Sample Overlap Evidence (CN)"),
            ]
            for r in top_pair_rows[:100]
        ],
    )
    make_sheet(
        out_wb,
        "Anchors",
        ["Anchor", "Type", "Candidate Rows", "Physical Pair Count"],
        [[a, anchor_type(a), c, anchor_pair_counts[a]] for a, c in anchor_counts.most_common()],
    )
    make_sheet(
        out_wb,
        "Manual vs Physical",
        ["Category", "Person", "PBSC"],
        [["manual_and_physical", p, pb] for p, pb in manual_with_physical]
        + [["manual_only", p, pb] for p, pb in manual_without_physical]
        + [["physical_only", p, pb] for p, pb in physical_without_manual[:250]],
    )
    out_wb.save(OUT_XLSX)

    def top_list(items, n=5):
        return "; ".join(f"{a} ({c})" for a, c in items[:n]) or "none"

    lines = []
    lines.append("# PBSC Overlap Analysis")
    lines.append("")
    lines.append("## Headline")
    lines.append(f"- {summary['central_committee_rows']} Central Committee rows, {summary['baidu_episode_rows']} Baidu career episodes, and {summary['baidu_gap_rows']} Baidu gap-review rows.")
    lines.append(f"- {summary['physical_candidate_rows']} strict physical-overlap candidate rows involving {summary['unique_people_with_physical_overlap']} non-PBSC people.")
    lines.append(f"- {summary['manual_assignments']} manual PBSC faction assignments imported from the faction workbook.")
    lines.append("")
    lines.append("## PBSC Profiles")
    for pbsc, info in by_pbsc.items():
        lines.append(f"- {pbsc} ({info['pbsc_chinese']}): {info['physical_people']} people with physical overlap, {info['manual_people']} manual labels, {info['physical_pairs']} physical episode pairs. Top places: {top_list(info['top_places'])}. Top institutions/orgs: {top_list(info['top_institutions'])}.")
    lines.append("")
    lines.append("## Strongest Physical Connectors")
    for r in top_connectors[:15]:
        lines.append(f"- {r['person']} / {r['pinyin']} ({r['title']}): overlaps {r['pbsc_count']} PBSC members; score {r['score']}; anchors {r['anchors']}; PBSCs {r['pbscs']}.")
    lines.append("")
    lines.append("## Strongest Pair-Level Evidence")
    for r in top_pair_rows[:15]:
        lines.append(
            f"- {r.get('Person Chinese Name')} -> {r.get('PBSC Pinyin Name')}: score {r.get('Strict Overlap Score')}, pairs {r.get('Physical Overlap Pairs')}, anchors {r.get('Shared Physical Anchors')}, years {r.get('Earliest Overlap Start')}-{r.get('Latest Overlap End')}."
        )
    lines.append("")
    lines.append("## Anchor Concentration")
    for anchor, count in anchor_counts.most_common(20):
        lines.append(f"- {anchor}: {count} candidate rows, {anchor_pair_counts[anchor]} physical episode pairs.")
    lines.append("")
    lines.append("## Manual-vs-Physical Notes")
    lines.append(f"- Manual labels with physical overlap in the strict data: {len(manual_with_physical)}.")
    lines.append(f"- Manual labels without detected physical overlap: {len(manual_without_physical)}.")
    lines.append(f"- Physical overlaps not in the manual faction workbook: {len(physical_without_manual)}.")
    OUT_MD.write_text("\n".join(lines), encoding="utf-8")

    print(json.dumps({
        "summary": summary,
        "top_pbsc": by_pbsc,
        "top_connectors": top_connectors[:15],
        "top_pairs": [
            {
                "person": r.get("Person Chinese Name"),
                "pinyin": r.get("Person Pinyin Name"),
                "pbsc": r.get("PBSC Pinyin Name"),
                "score": r.get("Strict Overlap Score"),
                "pairs": r.get("Physical Overlap Pairs"),
                "anchors": r.get("Shared Physical Anchors"),
                "years": f"{r.get('Earliest Overlap Start')}-{r.get('Latest Overlap End')}",
            }
            for r in top_pair_rows[:20]
        ],
        "top_anchors": anchor_counts.most_common(20),
        "manual_with_physical": len(manual_with_physical),
        "manual_without_physical": len(manual_without_physical),
        "physical_without_manual": len(physical_without_manual),
        "outputs": [str(OUT_MD), str(OUT_XLSX)],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
