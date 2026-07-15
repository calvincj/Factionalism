import csv
import math
import re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

from enrich_baidu import (
    CURRENT_YEAR,
    KNOWN_PBSC_CN,
    WORKBOOK,
    anchor_interval,
    autosize,
    intervals_overlap,
    overlap_window,
    split_tags,
    year_interval,
)


ROOT = Path(r"C:\Users\AMatthias\Documents\Codex\2026-05-12\i-want-to-make-database-that")
OUT_XLSX = ROOT / "PBSC visualization dataset.xlsx"
OUT_EDGE_CSV = ROOT / "PBSC visualization edges.csv"
OUT_SUMMARY_CSV = ROOT / "PBSC visualization person-pbsc summary.csv"

EXCLUDED_ANCHORS = {"中央党校", "中共中央党校"}
PBSC_START_YEAR = 2022

PROVINCES = {
    "北京", "天津", "上海", "重庆", "河北", "山西", "辽宁", "吉林", "黑龙江", "江苏", "浙江", "安徽",
    "福建", "江西", "山东", "河南", "湖北", "湖南", "广东", "海南", "四川", "贵州", "云南", "陕西",
    "甘肃", "青海", "台湾", "内蒙古", "广西", "西藏", "宁夏", "新疆", "香港", "澳门",
}


def rows_as_dicts(ws):
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(v or "").strip() for v in rows[0]]
    return [
        {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        for values in rows[1:]
        if any(v not in (None, "") for v in values)
    ]


def clean(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def years_label(start, end):
    return f"{start}-{end or CURRENT_YEAR}"


def classify_anchor(anchor):
    if anchor in PROVINCES:
        return "province/municipality"
    if any(keyword in anchor for keyword in ["大学", "学院", "科学院", "工程院", "中国社会科学院", "北京大学", "清华大学", "复旦大学", "浙江大学"]):
        return "university/research"
    if anchor.startswith("中央") or anchor in {"中共中央办公厅", "中央办公厅", "中央组织部", "中央宣传部", "中央统战部", "中央政法委", "中央纪委", "中央政策研究室"}:
        return "central party organ"
    if any(keyword in anchor for keyword in ["部", "委", "署", "国务院", "最高人民法院", "最高人民检察院"]):
        return "state/central institution"
    if any(keyword in anchor for keyword in ["银行", "集团", "公司", "国企", "保险"]):
        return "SOE/finance"
    return "other institution"


def anchor_weight(anchor_type):
    return {
        "province/municipality": 1.5,
        "central party organ": 1.4,
        "state/central institution": 1.2,
        "SOE/finance": 0.9,
        "university/research": 0.6,
        "other institution": 0.8,
    }.get(anchor_type, 0.8)


def role_rank(role):
    role = clean(role)
    if not role:
        return 0, "unknown"
    if "学习" in role and not any(k in role for k in ["书记", "部长", "主任", "主席", "省长", "市长", "委员", "副"]):
        return 1, "education/training"
    if any(k in role for k in ["总书记", "中央政治局常委", "国家主席", "中央军委主席"]):
        return 6, "top leadership/PBSC"
    if any(k in role for k in ["中央政治局委员", "国务院副总理", "国务委员", "中央书记处书记"]):
        return 5, "national leadership"
    if any(k in role for k in ["省委书记", "自治区党委书记", "直辖市委书记", "省长", "自治区主席", "部长", "最高人民法院院长", "最高人民检察院检察长", "全国政协副主席", "全国人大常委会副委员长"]):
        return 4, "ministerial/provincial chief"
    if any(k in role for k in ["省委常委", "自治区党委常委", "副省长", "副部长", "副主席", "副书记", "副主任", "秘书长"]):
        return 3, "vice-ministerial/provincial deputy"
    if any(k in role for k in ["市委书记", "市长", "厅长", "司长", "处长", "局长", "党委书记", "总经理", "董事长"]):
        return 2, "local/department/enterprise leadership"
    if any(k in role for k in ["干部", "教师", "研究人员", "工作人员", "科员"]):
        return 1, "early career/staff"
    return 2, "other official role"


def rank_stage(rank):
    return {
        0: "unknown",
        1: "education/early career",
        2: "local/department leadership",
        3: "vice-ministerial/provincial deputy",
        4: "ministerial/provincial chief",
        5: "national leadership",
        6: "top leadership/PBSC",
    }.get(rank, "unknown")


def relationship_label(person_rank, pbsc_rank):
    if not person_rank or not pbsc_rank:
        return "unknown"
    diff = pbsc_rank - person_rank
    if diff >= 2:
        return "PBSC member clearly senior at overlap"
    if diff == 1:
        return "PBSC member likely senior at overlap"
    if diff == 0:
        return "near-peer rank at overlap"
    return "connected individual senior/unclear at overlap"


def merge_windows(windows):
    if not windows:
        return []
    ordered = sorted((int(s), int(e)) for s, e in windows)
    merged = [list(ordered[0])]
    for start, end in ordered[1:]:
        if start <= merged[-1][1] + 1:
            merged[-1][1] = max(merged[-1][1], end)
        else:
            merged.append([start, end])
    return [(s, e) for s, e in merged]


def union_years(windows):
    return sum(end - start + 1 for start, end in merge_windows(windows))


def confidence(anchor_type_name, total_years, pair_count, relationship):
    if anchor_type_name == "province/municipality":
        if total_years >= 5 or pair_count >= 5:
            return "High", "same province/municipality with sustained or repeated overlapping career episodes"
        return "Medium", "same province/municipality with limited overlapping years"
    if anchor_type_name in {"central party organ", "state/central institution"}:
        if total_years >= 3 or pair_count >= 3:
            return "High", "same named central organization with sustained or repeated overlap"
        return "Medium", "same named central organization with limited overlap"
    if anchor_type_name == "SOE/finance":
        if total_years >= 3:
            return "Medium", "same named SOE/financial institution with sustained overlap"
        return "Low", "same SOE/financial anchor, but limited overlap"
    if anchor_type_name == "university/research":
        if total_years >= 3 and "senior" in relationship:
            return "Medium", "same university/research anchor plus seniority relationship"
        return "Low", "same university/research anchor is weaker evidence of factional connection"
    return "Low", "same named institution, but substantive relationship is unclear"


def timing_bucket(latest_end):
    if latest_end >= PBSC_START_YEAR:
        return "active/after 20th Congress"
    years = PBSC_START_YEAR - latest_end
    if years <= 5:
        return "recent before 20th Congress"
    if years <= 10:
        return "medium-term before 20th Congress"
    return "older overlap"


def data_quality(status, episode_count):
    status = clean(status)
    try:
        count = int(episode_count or 0)
    except ValueError:
        count = 0
    if status != "matched_title" or count == 0:
        return "gap/needs fallback"
    if count < 8:
        return "partial"
    return "usable"


def episode_record(row):
    interval = year_interval(row["Start Year"], row["End Year"])
    anchors = []
    for anchor in split_tags(row["Detected Places"]):
        if anchor in EXCLUDED_ANCHORS:
            continue
        anchors.append({
            "anchor": anchor,
            "type": classify_anchor(anchor),
            "interval": anchor_interval(anchor, row["Raw Timeline Entry (CN)"], interval),
        })
    rank, stage = role_rank(row["Role Text (CN)"])
    return {
        "pinyin": row["Pinyin Name"],
        "name": clean(row["Chinese Name"]),
        "is_pbsc": row["Is 20th PBSC"] == "Y",
        "episode_no": row["Episode No."],
        "interval": interval,
        "start": row["Start Year"],
        "end": row["End Year"],
        "role": clean(row["Role Text (CN)"]),
        "raw": clean(row["Raw Timeline Entry (CN)"]),
        "systems": set(split_tags(row["Detected System Tags"])),
        "anchors": anchors,
        "rank": rank,
        "stage": stage,
    }


def main():
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    main_rows = rows_as_dicts(wb["20th Central Committee"])
    episode_rows = rows_as_dicts(wb["Baidu Career Episodes"])
    faction_rows = rows_as_dicts(wb["PBSC Faction Assignments"])
    gap_rows = rows_as_dicts(wb["Baidu Gap Review"])

    title_by_name = {clean(r["Chinese Name"]): r.get("Title") for r in main_rows if r.get("Chinese Name")}
    branch_by_name = {clean(r["Chinese Name"]): r.get("Branch of government") for r in main_rows if r.get("Chinese Name")}
    status_by_name = {clean(r["Chinese Name"]): r.get("Baidu Match Status") for r in main_rows if r.get("Chinese Name")}
    episode_count_by_name = {clean(r["Chinese Name"]): r.get("Baidu Timeline Episode Count") for r in main_rows if r.get("Chinese Name")}
    gap_names = {clean(r["Chinese Name"]) for r in gap_rows if r.get("Chinese Name")}
    all_people = [(clean(r["Chinese Name"]), r["Pinyin Name"]) for r in main_rows if r.get("Chinese Name")]
    non_pbsc_people = [(cn, py) for cn, py in all_people if cn not in KNOWN_PBSC_CN]

    manual_by_pair = defaultdict(list)
    for row in faction_rows:
        manual_by_pair[(clean(row["Person Chinese Name"]), row["PBSC Pinyin Name"])].append(clean(row["Faction/Overlap Note"]))

    episodes = [episode_record(row) for row in episode_rows]
    episodes = [ep for ep in episodes if ep["interval"]]
    pbsc_eps = [ep for ep in episodes if ep["is_pbsc"]]
    person_eps = [ep for ep in episodes if not ep["is_pbsc"]]

    edge_groups = defaultdict(lambda: {
        "pair_count": 0,
        "windows": [],
        "evidence": [],
        "systems": Counter(),
        "person_ranks": [],
        "pbsc_ranks": [],
        "person_stages": Counter(),
        "pbsc_stages": Counter(),
    })

    for person_name, person_pinyin in non_pbsc_people:
        left_eps = [ep for ep in person_eps if ep["name"] == person_name]
        for pbsc_name in sorted(KNOWN_PBSC_CN):
            right_eps = [ep for ep in pbsc_eps if ep["name"] == pbsc_name]
            if not right_eps:
                continue
            pbsc_pinyin = right_eps[0]["pinyin"]
            for left in left_eps:
                for right in right_eps:
                    if not intervals_overlap(left["interval"], right["interval"]):
                        continue
                    for left_anchor in left["anchors"]:
                        for right_anchor in right["anchors"]:
                            if left_anchor["anchor"] != right_anchor["anchor"]:
                                continue
                            if not intervals_overlap(left_anchor["interval"], right_anchor["interval"]):
                                continue
                            start, end = overlap_window(left_anchor["interval"], right_anchor["interval"])
                            key = (person_name, person_pinyin, pbsc_name, pbsc_pinyin, left_anchor["anchor"])
                            rec = edge_groups[key]
                            rec["pair_count"] += 1
                            rec["windows"].append((start, end))
                            rec["systems"].update(left["systems"] & right["systems"])
                            rec["person_ranks"].append(left["rank"])
                            rec["pbsc_ranks"].append(right["rank"])
                            rec["person_stages"][left["stage"]] += 1
                            rec["pbsc_stages"][right["stage"]] += 1
                            if len(rec["evidence"]) < 5:
                                rec["evidence"].append(
                                    f"{left_anchor['anchor']} {start}-{end}: {years_label(left['start'], left['end'])} {left['role']} ↔ {pbsc_name} {years_label(right['start'], right['end'])} {right['role']}"
                                )

    edge_rows = []
    for key, rec in edge_groups.items():
        person_name, person_pinyin, pbsc_name, pbsc_pinyin, anchor = key
        merged = merge_windows(rec["windows"])
        total_years = union_years(rec["windows"])
        max_continuous = max((end - start + 1 for start, end in merged), default=0)
        earliest = min((start for start, _ in merged), default="")
        latest = max((end for _, end in merged), default="")
        anchor_type_name = classify_anchor(anchor)
        person_rank = max(rec["person_ranks"] or [0])
        pbsc_rank = max(rec["pbsc_ranks"] or [0])
        relationship = relationship_label(person_rank, pbsc_rank)
        conf, rationale = confidence(anchor_type_name, total_years, rec["pair_count"], relationship)
        manual_notes = manual_by_pair.get((person_name, pbsc_pinyin), [])
        weighted_score = round(
            (total_years * anchor_weight(anchor_type_name))
            + (rec["pair_count"] * 0.25)
            + (5 if manual_notes else 0)
            + (2 if conf == "High" else 1 if conf == "Medium" else 0),
            2,
        )
        edge_rows.append({
            "Person Chinese Name": person_name,
            "Person Pinyin Name": person_pinyin,
            "Person Title": title_by_name.get(person_name, ""),
            "Person Branch": branch_by_name.get(person_name, ""),
            "PBSC Chinese Name": pbsc_name,
            "PBSC Pinyin Name": pbsc_pinyin,
            "Shared Anchor": anchor,
            "Anchor Type": anchor_type_name,
            "Binary Overlap": 1,
            "Episode Pair Count": rec["pair_count"],
            "Total Overlap Years": total_years,
            "Max Continuous Overlap Years": max_continuous,
            "Earliest Overlap Start": earliest,
            "Latest Overlap End": latest,
            "Years From Last Overlap To 20th CC": max(0, PBSC_START_YEAR - int(latest)) if latest != "" else "",
            "Timing Bucket": timing_bucket(int(latest)) if latest != "" else "",
            "Person Rank At Overlap": person_rank,
            "Person Career Stage At Overlap": rank_stage(person_rank),
            "PBSC Member Rank At Overlap": pbsc_rank,
            "PBSC Member Career Stage At Overlap": rank_stage(pbsc_rank),
            "Role Relationship": relationship,
            "Context System Tags": "; ".join(sorted(rec["systems"])),
            "Mechanical Confidence": conf,
            "Confidence Rationale": rationale,
            "Manual Faction Label": "Y" if manual_notes else "",
            "Manual Faction Notes": "; ".join(manual_notes),
            "Weighted Overlap Score": weighted_score,
            "Person Biography Status": status_by_name.get(person_name, ""),
            "Person Episode Count": episode_count_by_name.get(person_name, ""),
            "Person Data Quality": data_quality(status_by_name.get(person_name), episode_count_by_name.get(person_name)),
            "Person Is Baidu Gap": "Y" if person_name in gap_names else "",
            "Sample Evidence": "\n".join(rec["evidence"]),
        })

    edge_rows.sort(key=lambda r: (r["Person Chinese Name"], r["PBSC Pinyin Name"], r["Shared Anchor"]))

    summary_groups = defaultdict(lambda: {
        "anchors": set(),
        "anchor_types": set(),
        "pair_count": 0,
        "total_years": 0,
        "max_continuous": 0,
        "weighted_score": 0.0,
        "confidences": Counter(),
        "manual": "",
        "manual_notes": set(),
        "earliest": None,
        "latest": None,
    })
    for row in edge_rows:
        key = (row["Person Chinese Name"], row["Person Pinyin Name"], row["PBSC Chinese Name"], row["PBSC Pinyin Name"])
        rec = summary_groups[key]
        rec["anchors"].add(row["Shared Anchor"])
        rec["anchor_types"].add(row["Anchor Type"])
        rec["pair_count"] += int(row["Episode Pair Count"])
        rec["total_years"] += int(row["Total Overlap Years"])
        rec["max_continuous"] = max(rec["max_continuous"], int(row["Max Continuous Overlap Years"]))
        rec["weighted_score"] += float(row["Weighted Overlap Score"])
        rec["confidences"][row["Mechanical Confidence"]] += 1
        if row["Manual Faction Label"] == "Y":
            rec["manual"] = "Y"
            rec["manual_notes"].add(row["Manual Faction Notes"])
        start = row["Earliest Overlap Start"]
        end = row["Latest Overlap End"]
        if start != "":
            rec["earliest"] = start if rec["earliest"] is None else min(rec["earliest"], start)
        if end != "":
            rec["latest"] = end if rec["latest"] is None else max(rec["latest"], end)

    summary_rows = []
    for key, rec in summary_groups.items():
        person_name, person_pinyin, pbsc_name, pbsc_pinyin = key
        strongest_conf = "High" if rec["confidences"]["High"] else "Medium" if rec["confidences"]["Medium"] else "Low"
        summary_rows.append({
            "Person Chinese Name": person_name,
            "Person Pinyin Name": person_pinyin,
            "Person Title": title_by_name.get(person_name, ""),
            "PBSC Chinese Name": pbsc_name,
            "PBSC Pinyin Name": pbsc_pinyin,
            "Binary Overlap": 1,
            "Distinct Anchor Count": len(rec["anchors"]),
            "Shared Anchors": "; ".join(sorted(rec["anchors"])),
            "Anchor Types": "; ".join(sorted(rec["anchor_types"])),
            "Total Episode Pair Count": rec["pair_count"],
            "Total Anchor-Years": rec["total_years"],
            "Max Continuous Overlap Years": rec["max_continuous"],
            "Earliest Overlap Start": rec["earliest"],
            "Latest Overlap End": rec["latest"],
            "Years From Last Overlap To 20th CC": max(0, PBSC_START_YEAR - int(rec["latest"])) if rec["latest"] is not None else "",
            "Timing Bucket": timing_bucket(int(rec["latest"])) if rec["latest"] is not None else "",
            "Strongest Mechanical Confidence": strongest_conf,
            "Manual Faction Label": rec["manual"],
            "Manual Faction Notes": "; ".join(sorted(filter(None, rec["manual_notes"]))),
            "Weighted Overlap Score": round(rec["weighted_score"], 2),
            "Person Data Quality": data_quality(status_by_name.get(person_name), episode_count_by_name.get(person_name)),
        })
    summary_rows.sort(key=lambda r: (-r["Weighted Overlap Score"], r["Person Chinese Name"], r["PBSC Pinyin Name"]))

    people_with_edges = {row["Person Chinese Name"] for row in edge_rows}
    person_summary = []
    for person_name, person_pinyin in non_pbsc_people:
        rows = [row for row in summary_rows if row["Person Chinese Name"] == person_name]
        person_summary.append({
            "Person Chinese Name": person_name,
            "Person Pinyin Name": person_pinyin,
            "Person Title": title_by_name.get(person_name, ""),
            "Has Strict Physical Overlap": "Y" if rows else "",
            "PBSC Count": len({r["PBSC Pinyin Name"] for r in rows}),
            "Total Weighted Score": round(sum(float(r["Weighted Overlap Score"]) for r in rows), 2),
            "Total Anchor-Years": sum(int(r["Total Anchor-Years"]) for r in rows),
            "Distinct Anchors": "; ".join(sorted({a for r in rows for a in split_tags(r["Shared Anchors"])})),
            "Connected PBSCs": "; ".join(sorted({r["PBSC Pinyin Name"] for r in rows})),
            "Manual Faction Matches": "; ".join(sorted({pb for (name, pb), notes in manual_by_pair.items() if name == person_name})),
            "Person Data Quality": data_quality(status_by_name.get(person_name), episode_count_by_name.get(person_name)),
            "Person Is Baidu Gap": "Y" if person_name in gap_names else "",
            "Baidu Episode Count": episode_count_by_name.get(person_name, ""),
        })
    person_summary.sort(key=lambda r: (-int(r["PBSC Count"]), -float(r["Total Weighted Score"]), r["Person Chinese Name"]))

    pbsc_location = defaultdict(lambda: {"people": set(), "pairs": 0, "years": 0, "score": 0.0, "top": Counter()})
    for row in edge_rows:
        key = (row["PBSC Chinese Name"], row["PBSC Pinyin Name"], row["Shared Anchor"], row["Anchor Type"])
        rec = pbsc_location[key]
        rec["people"].add(row["Person Chinese Name"])
        rec["pairs"] += int(row["Episode Pair Count"])
        rec["years"] += int(row["Total Overlap Years"])
        rec["score"] += float(row["Weighted Overlap Score"])
        rec["top"][f"{row['Person Chinese Name']} / {row['Person Pinyin Name']} - {row['Person Title']}"] += int(row["Episode Pair Count"])

    pbsc_location_rows = []
    for key, rec in pbsc_location.items():
        pbsc_name, pbsc_pinyin, anchor, anchor_type_name = key
        pbsc_location_rows.append({
            "PBSC Chinese Name": pbsc_name,
            "PBSC Pinyin Name": pbsc_pinyin,
            "Shared Anchor": anchor,
            "Anchor Type": anchor_type_name,
            "Connected People Count": len(rec["people"]),
            "Episode Pair Count": rec["pairs"],
            "Total Anchor-Years": rec["years"],
            "Weighted Score": round(rec["score"], 2),
            "Top Connected People": "; ".join(f"{person} ({count})" for person, count in rec["top"].most_common(8)),
        })
    pbsc_location_rows.sort(key=lambda r: (r["PBSC Pinyin Name"], -r["Weighted Score"], r["Shared Anchor"]))

    no_overlap_rows = [
        row for row in person_summary
        if row["Has Strict Physical Overlap"] != "Y"
    ]

    data_dictionary = [
        ["Binary Overlap", "1 when a person and PBSC member overlap at a specific non-excluded physical anchor."],
        ["Episode Pair Count", "Raw count of overlapping career episode pairs at this anchor. Useful but can be inflated by detailed timelines."],
        ["Total Overlap Years", "Unioned year count across all overlap windows for a person-PBSC-anchor row."],
        ["Max Continuous Overlap Years", "Longest continuous overlap window at the anchor."],
        ["Anchor Type", "Province/municipality, central party organ, state/central institution, SOE/finance, university/research, or other institution."],
        ["Mechanical Confidence", "Heuristic confidence based on anchor type, overlap duration, and pair count. It is not proof of factional alignment."],
        ["Role Relationship", "Heuristic comparison of the connected person's rank and the PBSC member's rank during overlap."],
        ["Timing Bucket", "How close the latest overlap is to the 20th Party Congress period."],
        ["Manual Faction Label", "Y if the person-PBSC pair appears in the separate faction workbook."],
        ["Person Data Quality", "Usable, partial, or gap/needs fallback based on Baidu match status and parsed episode count."],
    ]

    out = openpyxl.Workbook()
    out.remove(out.active)

    def write_sheet(name, rows):
        ws = out.create_sheet(name)
        if not rows:
            return ws
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h, "") for h in headers])
        ws.freeze_panes = "A2"
        autosize(ws, max_width=100)
        return ws

    write_sheet("Visualization Edges", edge_rows)
    write_sheet("Person PBSC Summary", summary_rows)
    write_sheet("Person Summary", person_summary)
    write_sheet("PBSC Location Summary", pbsc_location_rows)
    write_sheet("No Overlap Review", no_overlap_rows)
    dict_ws = out.create_sheet("Data Dictionary")
    dict_ws.append(["Field", "Definition"])
    for row in data_dictionary:
        dict_ws.append(row)
    autosize(dict_ws, max_width=100)
    out.save(OUT_XLSX)

    def write_csv(path, rows):
        if not rows:
            return
        with path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)

    write_csv(OUT_EDGE_CSV, edge_rows)
    write_csv(OUT_SUMMARY_CSV, summary_rows)

    print(OUT_XLSX)
    print(OUT_EDGE_CSV)
    print(OUT_SUMMARY_CSV)
    print(f"edges={len(edge_rows)} summaries={len(summary_rows)} people={len(person_summary)} no_overlap={len(no_overlap_rows)}")


if __name__ == "__main__":
    main()
