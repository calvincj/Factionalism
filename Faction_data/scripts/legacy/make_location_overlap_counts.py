from collections import defaultdict
from pathlib import Path

import openpyxl

from enrich_baidu import (
    CURRENT_YEAR,
    KNOWN_PBSC_CN,
    WORKBOOK,
    anchor_interval,
    autosize,
    intervals_overlap,
    overlap_window,
    split_tags,
    year_interval,
)


OUT = Path(r"C:\Users\AMatthias\Documents\Codex\2026-05-12\i-want-to-make-database-that\PBSC location overlap counts.xlsx")
SHEET_NAME = "Location Overlap Counts"
SUMMARY_SHEET = "Person Location Summary"
PBSC_SUMMARY_SHEET = "PBSC Location Summary"


def rows_as_dicts(ws):
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(v or "").strip() for v in rows[0]]
    return [
        {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        for values in rows[1:]
        if any(v not in (None, "") for v in values)
    ]


def episode_records(row):
    interval = year_interval(row["Start Year"], row["End Year"])
    anchors = split_tags(row["Detected Places"])
    anchor_records = []
    for anchor in anchors:
        if anchor in {"中央党校", "中共中央党校"}:
            continue
        anchor_records.append({
            "anchor": anchor,
            "interval": anchor_interval(anchor, row["Raw Timeline Entry (CN)"], interval),
        })
    return {
        "pinyin": row["Pinyin Name"],
        "name": row["Chinese Name"],
        "is_pbsc": row["Is 20th PBSC"] == "Y",
        "episode_no": row["Episode No."],
        "interval": interval,
        "start": row["Start Year"],
        "end": row["End Year"],
        "role": row["Role Text (CN)"],
        "raw": row["Raw Timeline Entry (CN)"],
        "anchors": anchor_records,
    }


def years(start, end):
    return f"{start}-{end or CURRENT_YEAR}"


def main():
    source = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    main_rows = rows_as_dicts(source["20th Central Committee"])
    title_by_name = {row["Chinese Name"]: row.get("Title") for row in main_rows if row.get("Chinese Name")}
    episode_rows = rows_as_dicts(source["Baidu Career Episodes"])
    episodes = [episode_records(row) for row in episode_rows]
    episodes = [ep for ep in episodes if ep["interval"]]

    pbsc_eps = [ep for ep in episodes if ep["is_pbsc"]]
    person_eps = [ep for ep in episodes if not ep["is_pbsc"]]
    people = sorted({(ep["name"], ep["pinyin"]) for ep in person_eps})

    counts = defaultdict(lambda: {
        "pair_count": 0,
        "earliest": None,
        "latest": None,
        "evidence": [],
    })

    for person_name, person_pinyin in people:
        left_eps = [ep for ep in person_eps if ep["name"] == person_name]
        for pbsc_name in sorted(KNOWN_PBSC_CN):
            right_eps = [ep for ep in pbsc_eps if ep["name"] == pbsc_name]
            if not right_eps:
                continue
            pbsc_pinyin = right_eps[0]["pinyin"]
            for left in left_eps:
                for right in right_eps:
                    if not intervals_overlap(left["interval"], right["interval"]):
                        continue
                    for left_anchor in left["anchors"]:
                        for right_anchor in right["anchors"]:
                            if left_anchor["anchor"] != right_anchor["anchor"]:
                                continue
                            if not intervals_overlap(left_anchor["interval"], right_anchor["interval"]):
                                continue
                            anchor = left_anchor["anchor"]
                            start, end = overlap_window(left_anchor["interval"], right_anchor["interval"])
                            key = (person_name, person_pinyin, pbsc_name, pbsc_pinyin, anchor)
                            rec = counts[key]
                            rec["pair_count"] += 1
                            rec["earliest"] = start if rec["earliest"] is None else min(rec["earliest"], start)
                            rec["latest"] = end if rec["latest"] is None else max(rec["latest"], end)
                            if len(rec["evidence"]) < 3:
                                rec["evidence"].append(
                                    f"{anchor} {start}-{end}: {years(left['start'], left['end'])} {left['role']} ↔ {pbsc_name} {years(right['start'], right['end'])} {right['role']}"
                                )

    out = openpyxl.Workbook()
    ws = out.active
    ws.title = SHEET_NAME
    ws.append([
        "Person Chinese Name",
        "Person Pinyin Name",
        "Person Title",
        "PBSC Chinese Name",
        "PBSC Pinyin Name",
        "Shared Location/Institution",
        "Overlap Pair Count",
        "Earliest Overlap Start",
        "Latest Overlap End",
        "Sample Evidence",
    ])
    for key, rec in sorted(counts.items(), key=lambda item: (item[0][0], item[0][3], -item[1]["pair_count"], item[0][4])):
        person_name, person_pinyin, pbsc_name, pbsc_pinyin, anchor = key
        ws.append([
            person_name,
            person_pinyin,
            title_by_name.get(person_name, ""),
            pbsc_name,
            pbsc_pinyin,
            anchor,
            rec["pair_count"],
            rec["earliest"],
            rec["latest"],
            "\n".join(rec["evidence"]),
        ])

    person_summary = out.create_sheet(SUMMARY_SHEET)
    person_summary.append([
        "Person Chinese Name",
        "Person Pinyin Name",
        "Person Title",
        "Total Location-Level Overlap Pairs",
        "PBSC Count",
        "Location/Institution Count",
        "Top Location Counts",
        "PBSCs",
    ])
    by_person = defaultdict(lambda: {"pairs": 0, "pbscs": set(), "anchors": defaultdict(int)})
    for key, rec in counts.items():
        person_name, person_pinyin, _pbsc_name, pbsc_pinyin, anchor = key
        item = by_person[(person_name, person_pinyin)]
        item["pairs"] += rec["pair_count"]
        item["pbscs"].add(pbsc_pinyin)
        item["anchors"][anchor] += rec["pair_count"]
    for (person_name, person_pinyin), rec in sorted(by_person.items(), key=lambda item: (-item[1]["pairs"], item[0][0])):
        top_anchors = "; ".join(f"{anchor} ({count})" for anchor, count in sorted(rec["anchors"].items(), key=lambda x: -x[1])[:8])
        person_summary.append([
            person_name,
            person_pinyin,
            title_by_name.get(person_name, ""),
            rec["pairs"],
            len(rec["pbscs"]),
            len(rec["anchors"]),
            top_anchors,
            "; ".join(sorted(rec["pbscs"])),
        ])

    pbsc_summary = out.create_sheet(PBSC_SUMMARY_SHEET)
    pbsc_summary.append([
        "PBSC Chinese Name",
        "PBSC Pinyin Name",
        "Shared Location/Institution",
        "People Count",
        "Overlap Pair Count",
        "Top People",
    ])
    by_pbsc_anchor = defaultdict(lambda: {"pairs": 0, "people": defaultdict(int)})
    for key, rec in counts.items():
        person_name, person_pinyin, pbsc_name, pbsc_pinyin, anchor = key
        item = by_pbsc_anchor[(pbsc_name, pbsc_pinyin, anchor)]
        item["pairs"] += rec["pair_count"]
        item["people"][f"{person_name} / {person_pinyin}"] += rec["pair_count"]
    for (pbsc_name, pbsc_pinyin, anchor), rec in sorted(by_pbsc_anchor.items(), key=lambda item: (item[0][1], -item[1]["pairs"])):
        top_people_parts = []
        for person, count in sorted(rec["people"].items(), key=lambda x: -x[1])[:8]:
            person_name = person.split(" / ", 1)[0]
            title = title_by_name.get(person_name, "")
            title_text = f" - {title}" if title else ""
            top_people_parts.append(f"{person}{title_text} ({count})")
        top_people = "; ".join(top_people_parts)
        pbsc_summary.append([
            pbsc_name,
            pbsc_pinyin,
            anchor,
            len(rec["people"]),
            rec["pairs"],
            top_people,
        ])

    for sheet in out.worksheets:
        sheet.freeze_panes = "A2"
        autosize(sheet, max_width=100)
    out.save(OUT)
    print(OUT)
    print(f"rows={ws.max_row - 1}")
    print(f"people={len(by_person)}")
    print(f"pbsc_location_rows={pbsc_summary.max_row - 1}")


if __name__ == "__main__":
    main()
