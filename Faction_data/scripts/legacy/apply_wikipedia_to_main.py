"""
Takes the main aggregated workbook (originally "PBSC overlap analysis tables -
with titles.xlsx", everything in it Baidu-derived) and:

1. Renames every column that holds Baidu-derived data so its source is explicit
   (e.g. "Score" -> "Baidu Score"), now that a second source exists.
2. Joins in Wikipedia cross-check results (scripts/enrich_wikipedia.py output)
   onto "Top Connectors" and "Top Pair Evidence" by person name.
3. Adds a new "Wikipedia Coverage (All 219)" sheet so people Wikipedia has data
   on but who never entered the Baidu candidate pool (0 Baidu episodes -> never
   generated an overlap-candidate row) become visible in the main file at all.
4. Adds Wikipedia coverage counts to the Summary sheet.

Does NOT change any score or ranking -- it only labels provenance and appends
data. The underlying Baidu-only overlap analysis is untouched.

Writes to outputs/PBSC Overlap Analysis - MAIN.xlsx (new name) and leaves the
old "with titles" file in place until you confirm the new one looks right.
"""

import re
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
OUTPUTS = SCRIPT_DIR.parent / "outputs"
SRC_XLSX = OUTPUTS / "PBSC overlap analysis tables - with titles.xlsx"
WIKI_XLSX = OUTPUTS / "PBSC Wikipedia cross-check.xlsx"
OUT_XLSX = OUTPUTS / "PBSC Overlap Analysis - MAIN.xlsx"

RENAME_MAP = {
    "By PBSC": {
        "Physical People": "Baidu Physical People",
        "Physical Pairs": "Baidu Physical Pairs",
        "Top Places": "Baidu Top Places",
        "Top Institutions": "Baidu Top Institutions",
    },
    "Top Connectors": {
        "Score": "Baidu Score",
        "Physical Pairs": "Baidu Physical Pairs",
        "Top Anchors": "Baidu Top Anchors",
    },
    "Top Pair Evidence": {
        "Score": "Baidu Score",
        "Pairs": "Baidu Pairs",
        "Anchors": "Baidu Anchors",
        "Evidence": "Baidu Evidence (CN)",
    },
    "Anchors": {
        "Candidate Rows": "Baidu Candidate Rows",
        "Physical Pair Count": "Baidu Physical Pair Count",
    },
}


def strip_paren(name):
    return re.sub(r"[（(].*?[）)]", "", str(name or "")).strip()


def autosize(ws, max_width=90):
    for column_cells in ws.columns:
        letter = column_cells[0].column_letter
        width = 10
        for cell in column_cells[:200]:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(max_width, len(value) + 2))
        ws.column_dimensions[letter].width = width


def load_wikipedia_lookup():
    wb = openpyxl.load_workbook(WIKI_XLSX, data_only=True)
    ws = wb["Wikipedia Cross-Check"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    lookup = {}
    all_rows = []
    for r in rows[1:]:
        d = dict(zip(headers, r))
        key = strip_paren(d["Chinese Name"])
        lookup[key] = d
        all_rows.append(d)
    wb.close()
    return lookup, all_rows, list(headers)


def sheet_to_rows(ws):
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    return headers, rows[1:]


def main():
    wiki_lookup, wiki_all_rows, wiki_headers = load_wikipedia_lookup()

    src_wb = openpyxl.load_workbook(SRC_XLSX, data_only=True)
    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    wiki_join_cols = ["Wikipedia ZH Status", "Wikipedia EN Status", "Wikipedia ZH Offices", "Wikipedia EN Offices"]

    for sheet_name in src_wb.sheetnames:
        src_ws = src_wb[sheet_name]
        headers, data_rows = sheet_to_rows(src_ws)
        rename = RENAME_MAP.get(sheet_name, {})
        new_headers = [rename.get(h, h) for h in headers]

        needs_join = sheet_name in ("Top Connectors", "Top Pair Evidence")
        person_idx = new_headers.index("Person") if needs_join and "Person" in new_headers else None

        out_ws = out_wb.create_sheet(sheet_name)
        out_ws.append(new_headers + (wiki_join_cols if needs_join else []))
        for row in data_rows:
            row = list(row)
            if needs_join:
                key = strip_paren(row[person_idx])
                w = wiki_lookup.get(key)
                if w:
                    row += [
                        w.get("Wikipedia ZH Status", ""),
                        w.get("Wikipedia EN Status", ""),
                        w.get("Wikipedia ZH Offices", ""),
                        w.get("Wikipedia EN Offices", ""),
                    ]
                else:
                    row += ["not_in_wikipedia_check", "not_in_wikipedia_check", "", ""]
            out_ws.append(row)
        out_ws.freeze_panes = "A2"
        autosize(out_ws)

    # New sheet: every one of the 219 people's Wikipedia status, including the
    # ones invisible in the Baidu-only sheets above (0 Baidu episodes -> never
    # generated a candidate row -> never appears in Top Connectors/Top Pair Evidence).
    coverage_ws = out_wb.create_sheet("Wikipedia Coverage (All 219)")
    coverage_headers = [
        "Chinese Name", "Pinyin Name", "Title", "Baidu Match Status", "Baidu Episode Count",
        "Wikipedia ZH Status", "Wikipedia ZH Offices", "Wikipedia EN Status", "Wikipedia EN Offices",
        "Cross-Check Flag",
    ]
    coverage_ws.append(coverage_headers)
    for d in wiki_all_rows:
        coverage_ws.append([d.get(h, "") for h in coverage_headers])
    coverage_ws.freeze_panes = "A2"
    autosize(coverage_ws)

    # Patch the Summary sheet with Wikipedia coverage counts.
    from collections import Counter
    flag_counts = Counter(d["Cross-Check Flag"] for d in wiki_all_rows)
    summary_ws = out_wb["Summary"]
    summary_ws.append(["wikipedia_people_processed", len(wiki_all_rows)])
    for flag, count in flag_counts.most_common():
        summary_ws.append([f"wikipedia_{flag}", count])
    autosize(summary_ws)

    out_wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")
    print(f"Sheets: {out_wb.sheetnames}")


if __name__ == "__main__":
    main()
